import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.workbook.defined_name import DefinedName

OUT = '/Users/dan/Documents/HomeServicesBrand/sorted_saas_recurring_revenue_forecast.xlsx'
wb = openpyxl.load_workbook(OUT)

FIRST = 5
BLOCK_W = 44
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
thin = Side(style='thin', color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '£#,##0'
PCT2 = '0.00%'

NAMES = ['Mo','CalMo','CalYear','AcqSeas','Infl','EffChurn','Mkt','PaidNew','RefNew','TotNew',
         'NewMo','NewAnn','NEsM','NPrM','NEsA','NPrA','AEsM','APrM','AEsA','APrA','TotCust',
         'MRR','ARR','MoBill','NewAnnCash','RenAnn','GrossRev','Roll12','VATReg','NetRev',
         'VATAccr','VATPay','VATBal','Prov','PayFee','OpsAI','FixedCost','RefCost','MktOut',
         'Profit','Cash','ARPU','Churn']
COL = {n: i for i, n in enumerate(NAMES, 1)}

SHARED = {
    'pessm':'PriceEssMonthly','pprmm':'PricePremMonthly','pessa':'PriceEssAnnual','pprma':'PricePremAnnual',
    'prov':'ProviderShare','pfee':'PaymentFee','ops':'OpsAI','fixed':'MonthlyFixed',
    'init':'InitialInvestment','buffer':'CashBuffer','refmix':'ReferralMix','refreward':'ReferralReward',
    'mkt13':'MktFixed1_3','mkt46':'MktFixed4_6','pct712':'MktPct7_12','pct1318':'MktPct13_18',
    'pct19':'MktPct19plus','min712':'MktMin7_12','min1318':'MktMin13_18','min19':'MktMin19plus',
    'essmix':'EssentialMix','pdelay':'ProviderDelay',
}

def cl(sc, n):
    return openpyxl.utils.get_column_letter(sc + COL[n] - 1)

def ch_expr(sc, r, ach):
    C = lambda n: cl(sc, n)
    return f"({ach}*VLOOKUP({C('CalMo')}{r},ChurnSeasonTable,4,FALSE))"

def monthly_active_scalar(sc, r, new, act):
    C = lambda n: cl(sc, n)
    m, a, f = C(new), C(act), C('EffChurn')
    if r == FIRST:
        return f"={m}{r}"
    if r == FIRST + 1:
        return f"={m}{r}+{m}{r-1}"
    if r == FIRST + 2:
        return f"={m}{r}+{m}{r-1}+{m}{r-2}"
    return f"={m}{r}+{m}{r-1}+{m}{r-2}+MAX(0,{a}{r-1}-{m}{r-1}-{m}{r-2})*(1-{f}{r})"

def anniversary_terms(sc, r, new, ach, template):
    C = lambda n: cl(sc, n)
    o = C(new)
    ch = ch_expr(sc, r, ach)
    parts = []
    for yrs, offset in [(0, 12), (1, 24), (2, 36), (3, 48)]:
        src = r - offset
        if src < FIRST:
            continue
        parts.append(template(o, src, ch, yrs))
    return parts

def annual_churn_sum(sc, r, new, ach):
    parts = anniversary_terms(
        sc, r, new, ach,
        lambda o, src, ch, yrs: f"{o}{src}*{ch}*POWER(1-{ch},{yrs})"
    )
    return "=0" if not parts else "=" + "+".join(parts)

def annual_active_scalar(sc, r, new, act, ach):
    C = lambda n: cl(sc, n)
    o, a = C(new), C(act)
    if r == FIRST:
        return f"={o}{r}"
    if r <= FIRST + 11:
        return f"=SUM({o}$5:{o}{r})"
    churn = annual_churn_sum(sc, r, new, ach).lstrip('=')
    return f"={a}{r-1}+{o}{r}-{churn}"

def ren_ann_scalar(sc, r, ach):
    C = lambda n: cl(sc, n)
    inf = C('Infl')

    def part(new, price):
        parts = anniversary_terms(
            sc, r, new, ach,
            lambda o, src, ch, yrs: f"{o}{src}*POWER(1-{ch},{yrs+1})*{price}*{inf}{r}"
        )
        return "+".join(parts) if parts else "0"

    return f"={part('NEsA', SHARED['pessa'])}+{part('NPrA', SHARED['pprma'])}"

def vat_pay(sc, r):
    B, C, AD, AB = cl(sc, 'CalMo'), cl(sc, 'CalYear'), cl(sc, 'VATAccr'), cl(sc, 'VATReg')
    cy, cm, va = f"${C}${FIRST}:${C}{r}", f"${B}${FIRST}:${B}{r}", f"${AD}${FIRST}:${AD}{r}"
    cr = f"${C}{r}"
    q1 = f"SUMPRODUCT(--(({cy}={cr}-1)*({cm}>=10)*({cm}<=12)),{va})"
    q2 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=1)*({cm}<=3)),{va})"
    q3 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=4)*({cm}<=6)),{va})"
    q4 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=7)*({cm}<=9)),{va})"
    return f"=IF({AB}{r}=0,0,IF({B}{r}=1,{q1},IF({B}{r}=4,{q2},IF({B}{r}=7,{q3},IF({B}{r}=10,{q4},0)))))"

def write_block(ws, sc, label, mchurn, achurn, annmix, cac, hdr_row=4):
    if label:
        ws.cell(2, sc, label).font = Font(bold=True, size=9, color="2F5496")
    hdrs = ["Mo","Cal","CY","Acq","Infl","ECh","Mkt","Paid","Ref","Tot","NMo","NAn",
            "NEsM","NPrM","NEsA","NPrA","AEsM","APrM","AEsA","APrA","Cust","MRR","ARR",
            "Mo$","NwAn","Ren","Gross","R12","VAT","Net","VAcc","VPay","VBal","Prov",
            "Fee","Ops","Fix","Ref$","Mkt","Pft","Cash","ARPU","Chrn"]
    for i, h in enumerate(hdrs):
        c = ws.cell(hdr_row, sc + i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    for m in range(1, 61):
        r = FIRST + m - 1
        C = lambda n: cl(sc, n)
        ws.cell(r, sc + COL['Mo'] - 1, m)
        ws.cell(r, sc + COL['CalMo'] - 1, f"=MOD({C('Mo')}{r}-1+LaunchMonth-1,12)+1")
        ws.cell(r, sc + COL['CalYear'] - 1, f"=INT(({C('Mo')}{r}-1+LaunchMonth-1)/12)+1")
        ws.cell(r, sc + COL['AcqSeas'] - 1, f"=VLOOKUP({C('CalMo')}{r},SeasonTable,2,FALSE)")
        ws.cell(r, sc + COL['Infl'] - 1, f"=(1+InflationRate)^INT(({C('Mo')}{r}-1)/12)")
        ws.cell(r, sc + COL['EffChurn'] - 1, f"={mchurn}*VLOOKUP({C('CalMo')}{r},ChurnSeasonTable,4,FALSE)")

        cb = f"{C('Cash')}{r-1}" if m > 1 else SHARED['init']
        rp = f"{C('GrossRev')}{r-1}" if m > 1 else "0"
        if m <= 3:
            mkt = f"={SHARED['mkt13']}"
        elif m <= 6:
            mkt = f"=MIN({SHARED['mkt46']},MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 12:
            mkt = f"=MIN(MAX({SHARED['min712']},{rp}*{SHARED['pct712']}),MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 18:
            mkt = f"=MIN(MAX({SHARED['min1318']},{rp}*{SHARED['pct1318']}),MAX(0,{cb}-{SHARED['buffer']}))"
        else:
            mkt = f"=MIN(MAX({SHARED['min19']},{rp}*{SHARED['pct19']}),MAX(0,{cb}-{SHARED['buffer']}))"
        ws.cell(r, sc + COL['Mkt'] - 1, mkt)
        ws.cell(r, sc + COL['PaidNew'] - 1, f"=IF({C('Mkt')}{r}>0,MAX(1,ROUND({C('Mkt')}{r}/{cac}*{C('AcqSeas')}{r},0)),0)")
        ws.cell(r, sc + COL['RefNew'] - 1, f"=IF({C('PaidNew')}{r}>0,ROUND({C('PaidNew')}{r}*{SHARED['refmix']}/(1-{SHARED['refmix']}),0),0)")
        ws.cell(r, sc + COL['TotNew'] - 1, f"={C('PaidNew')}{r}+{C('RefNew')}{r}")
        ws.cell(r, sc + COL['NewMo'] - 1, f"=ROUND({C('TotNew')}{r}*(1-{annmix}),0)")
        ws.cell(r, sc + COL['NewAnn'] - 1, f"={C('TotNew')}{r}-{C('NewMo')}{r}")
        ws.cell(r, sc + COL['NEsM'] - 1, f"=ROUND({C('NewMo')}{r}*{SHARED['essmix']},0)")
        ws.cell(r, sc + COL['NPrM'] - 1, f"={C('NewMo')}{r}-{C('NEsM')}{r}")
        ws.cell(r, sc + COL['NEsA'] - 1, f"=ROUND({C('NewAnn')}{r}*{SHARED['essmix']},0)")
        ws.cell(r, sc + COL['NPrA'] - 1, f"={C('NewAnn')}{r}-{C('NEsA')}{r}")
        ws.cell(r, sc + COL['AEsM'] - 1, monthly_active_scalar(sc, r, 'NEsM', 'AEsM'))
        ws.cell(r, sc + COL['APrM'] - 1, monthly_active_scalar(sc, r, 'NPrM', 'APrM'))
        ws.cell(r, sc + COL['AEsA'] - 1, annual_active_scalar(sc, r, 'NEsA', 'AEsA', achurn))
        ws.cell(r, sc + COL['APrA'] - 1, annual_active_scalar(sc, r, 'NPrA', 'APrA', achurn))

        inf = f"{C('Infl')}{r}"
        ws.cell(r, sc + COL['TotCust'] - 1, f"={C('AEsM')}{r}+{C('APrM')}{r}+{C('AEsA')}{r}+{C('APrA')}{r}")
        ws.cell(r, sc + COL['MRR'] - 1, f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{SHARED['pprmm']}*{inf}+{C('AEsA')}{r}*{SHARED['pessa']}*{inf}/12+{C('APrA')}{r}*{SHARED['pprma']}*{inf}/12")
        ws.cell(r, sc + COL['ARR'] - 1, f"={C('MRR')}{r}*12")
        ws.cell(r, sc + COL['MoBill'] - 1, f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{SHARED['pprmm']}*{inf}")
        ws.cell(r, sc + COL['NewAnnCash'] - 1, f"={C('NEsA')}{r}*{SHARED['pessa']}*{inf}+{C('NPrA')}{r}*{SHARED['pprma']}*{inf}")
        ws.cell(r, sc + COL['RenAnn'] - 1, ren_ann_scalar(sc, r, achurn))
        ws.cell(r, sc + COL['GrossRev'] - 1, f"={C('MoBill')}{r}+{C('NewAnnCash')}{r}+{C('RenAnn')}{r}")
        ws.cell(r, sc + COL['Roll12'] - 1, f"=SUM({C('GrossRev')}${FIRST}:{C('GrossRev')}{r})" if m < 12 else f"=SUM({C('GrossRev')}{r-11}:{C('GrossRev')}{r})")
        ws.cell(r, sc + COL['VATReg'] - 1, f"={C('Roll12')}{r}>=VATThreshold" if m == 1 else f"=OR({C('VATReg')}{r-1},{C('Roll12')}{r}>=VATThreshold)")
        ws.cell(r, sc + COL['NetRev'] - 1, f"=IF({C('VATReg')}{r},{C('GrossRev')}{r}/(1+VATRate),{C('GrossRev')}{r})")
        ws.cell(r, sc + COL['VATAccr'] - 1, f"={C('GrossRev')}{r}-{C('NetRev')}{r}")
        ws.cell(r, sc + COL['VATPay'] - 1, vat_pay(sc, r))
        ws.cell(r, sc + COL['VATBal'] - 1, f"={C('VATAccr')}{r}-{C('VATPay')}{r}" if m == 1 else f"={C('VATBal')}{r-1}+{C('VATAccr')}{r}-{C('VATPay')}{r}")

        if m == 1:
            prov = f"=IF({SHARED['pdelay']}>=1,0,{C('NetRev')}{r}*{SHARED['prov']})"
        else:
            prov = f"=IF({SHARED['pdelay']}>=1,{C('NetRev')}{r-1}*{SHARED['prov']},{C('NetRev')}{r}*{SHARED['prov']})"
        ws.cell(r, sc + COL['Prov'] - 1, prov)
        ws.cell(r, sc + COL['PayFee'] - 1, f"={C('GrossRev')}{r}*{SHARED['pfee']}")
        ws.cell(r, sc + COL['OpsAI'] - 1, f"={C('TotCust')}{r}*{SHARED['ops']}")
        ws.cell(r, sc + COL['FixedCost'] - 1, f"={SHARED['fixed']}")
        ws.cell(r, sc + COL['RefCost'] - 1, f"={C('RefNew')}{r}*{SHARED['refreward']}")
        ws.cell(r, sc + COL['MktOut'] - 1, f"={C('Mkt')}{r}")
        ws.cell(r, sc + COL['Profit'] - 1, f"={C('NetRev')}{r}-{C('Prov')}{r}-{C('PayFee')}{r}-{C('OpsAI')}{r}-{C('FixedCost')}{r}-{C('RefCost')}{r}-{C('MktOut')}{r}")
        out = f"{C('Prov')}{r}+{C('PayFee')}{r}+{C('OpsAI')}{r}+{C('FixedCost')}{r}+{C('RefCost')}{r}+{C('MktOut')}{r}+{C('VATPay')}{r}"
        ws.cell(r, sc + COL['Cash'] - 1, f"={SHARED['init']}+{C('GrossRev')}{r}-({out})" if m == 1 else f"={C('Cash')}{r-1}+{C('GrossRev')}{r}-({out})")
        ws.cell(r, sc + COL['ARPU'] - 1, f"=IF({C('TotCust')}{r}>0,{C('MRR')}{r}/{C('TotCust')}{r},0)")
        ws.cell(r, sc + COL['Churn'] - 1, 0 if m == 1 else f"=IF({C('TotCust')}{r-1}>0,MAX(0,({C('TotCust')}{r-1}+{C('TotNew')}{r}-{C('TotCust')}{r})/{C('TotCust')}{r-1}),0)")

        for ci in range(sc, sc + len(NAMES)):
            ws.cell(r, ci).border = BORDER
            ncol = ci - sc + 1
            if ncol in (COL['MRR'], COL['ARR'], COL['GrossRev'], COL['NetRev'], COL['Profit'], COL['Cash']):
                ws.cell(r, ci).number_format = MONEY
            elif ncol in (COL['Churn'], COL['EffChurn']):
                ws.cell(r, ci).number_format = PCT2

    return {k: cl(sc, k) for k in ['TotCust', 'MRR', 'ARR', 'GrossRev', 'Profit', 'Cash', 'PaidNew', 'RefNew', 'NetRev', 'VATPay']}


idx = wb.sheetnames.index('Model')
del wb['Model']
ws_m = wb.create_sheet('Model', idx)
ws_m['A1'] = "60-Month Model (scalar formulas - Excel compatible)"
ws_m['A1'].font = TITLE_FONT
M = write_block(ws_m, 1, None, 'MonthlyChurn', 'AnnualChurn', 'AnnualMix', 'CAC')

for m, lb in [(12, 'Month12'), (24, 'Month24'), (60, 'Month60')]:
    er = FIRST + m - 1
    for nm, cn in [('MRR', 'MRR'), ('ARR', 'ARR'), ('Cust', 'TotCust'), ('Cash', 'Cash')]:
        name = f"{nm}_{lb}"
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name, attr_text=f"Model!${M[cn]}${er}"))

if 'Scenario Engine' in wb.sheetnames:
    del wb['Scenario Engine']
ws_se = wb.create_sheet('Scenario Engine')
ws_se.sheet_state = 'hidden'

blocks = {}
blocks['cons'] = write_block(ws_se, 1, 'Conservative', "'Sens - Scenarios'!$B$5", "'Sens - Scenarios'!$B$6", "'Sens - Scenarios'!$B$7", "'Sens - Scenarios'!$B$8", hdr_row=3)
blocks['base'] = write_block(ws_se, 1 + BLOCK_W, 'Base', 'MonthlyChurn', 'AnnualChurn', 'AnnualMix', 'CAC', hdr_row=3)
blocks['agg'] = write_block(ws_se, 1 + BLOCK_W * 2, 'Aggressive', "'Sens - Scenarios'!$D$5", "'Sens - Scenarios'!$D$6", "'Sens - Scenarios'!$D$7", "'Sens - Scenarios'!$D$8", hdr_row=3)
for i in range(9):
    blocks[f'ch{i}'] = write_block(ws_se, 1 + BLOCK_W * (3 + i), f'Ch{i}', f"'Sens - Churn'!$A${13 + i}", 'AnnualChurn', 'AnnualMix', 'CAC', hdr_row=3)
for i in range(7):
    blocks[f'mx{i}'] = write_block(ws_se, 1 + BLOCK_W * (12 + i), f'Mx{i}', 'MonthlyChurn', 'AnnualChurn', f"'Sens - Annual Mix'!$A${9 + i}", 'CAC', hdr_row=3)
for i in range(7):
    blocks[f'cac{i}'] = write_block(ws_se, 1 + BLOCK_W * (19 + i), f'Cac{i}', 'MonthlyChurn', 'AnnualChurn', 'AnnualMix', f"'Sens - CAC & Marketing'!$A${9 + i}", hdr_row=3)

bad = []
for row in ws_m.iter_rows():
    for c in row:
        v = c.value
        if isinstance(v, str) and any(f'{col}-' in v for col in 'MNOPQRST'):
            bad.append((c.coordinate, v[:100]))

print(f"Bad refs: {len(bad)}")
print("S17:", ws_m.cell(17, 19).value)
print("Z17:", ws_m.cell(17, 26).value)
print("Z29:", ws_m.cell(29, 26).value)
n = sum(1 for row in ws_m.iter_rows() for c in row if isinstance(c.value, str) and 'SUMPRODUCT' in c.value)
print(f"SUMPRODUCT count: {n}")

wb.save(OUT)
print("Saved")
