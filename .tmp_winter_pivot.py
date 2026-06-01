"""Add Winter Pivot scenario: seasonality tables, Scenario Engine block, Sens tab."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

OUT = '/Users/dan/Documents/HomeServicesBrand/sorted_saas_recurring_revenue_forecast.xlsx'
wb = openpyxl.load_workbook(OUT)

FIRST = 5
BLOCK_W = 44
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, color="666666", size=10)
thin = Side(style='thin', color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '£#,##0'
PCT2 = '0.00%'

NAMES = ['Mo','CalMo','CalYear','AcqSeas','Infl','EffChurn','Mkt','PaidNew','RefNew','TotNew',
         'NewMo','NewAnn','NEsM','NPrM','NEsA','NPrA','AEsM','APrM','AEsA','APrA','TotCust',
         'MRR','ARR','MoBill','NewAnnCash','RenAnn','GrossRev','Roll12','VATReg','NetRev',
         'VATAccr','VATPay','VATBal','Prov','PayFee','OpsAI','FixedCost','RefCost','MktOut',
         'Profit','Cash','ARPU','Churn']
COL = {n: i for i, n in enumerate(NAMES, 1)}

SHARED = {
    'pessm':'PriceEssMonthly','pprmm':'PricePremMonthly','pessa':'PriceEssAnnual','pprma':'PricePremAnnual',
    'prov':'ProviderShare','pfee':'PaymentFee','ops':'OpsAI','fixed':'MonthlyFixed',
    'init':'InitialInvestment','buffer':'CashBuffer','refmix':'ReferralMix','refreward':'ReferralReward',
    'mkt13':'MktFixed1_3','mkt46':'MktFixed4_6','pct712':'MktPct7_12','pct1318':'MktPct13_18',
    'pct19':'MktPct19plus','min712':'MktMin7_12','min1318':'MktMin13_18','min19':'MktMin19plus',
    'pdelay':'ProviderDelay',
}

# Base acq (B) and churn mult (D) -> winter pivot acq (E) and churn mult (F)
# Oct–Feb adjusted for gutter / leaf / patio property-care bundle
WINTER_PIVOT = {
    1:  (0.72, 1.08),   # Jan
    2:  (0.82, 1.05),   # Feb
    3:  (0.87, 1.04),   # Mar - unchanged
    4:  (1.07, 0.85),
    5:  (1.17, 0.76),
    6:  (1.22, 0.71),
    7:  (1.22, 0.71),
    8:  (1.17, 0.76),
    9:  (1.02, 0.90),
    10: (1.00, 0.98),   # Oct - leaf/gutter season
    11: (1.05, 0.95),   # Nov - peak gutter/leaf
    12: (0.78, 1.08),   # Dec
}

MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

# --- Inputs: add winter pivot columns E & F ---
ws_in = wb['Inputs']
ws_in['E54'] = 'Pivot acq'
ws_in['F54'] = 'Pivot churn'
ws_in['E54'].font = Font(bold=True)
ws_in['F54'].font = Font(bold=True)

ws_in['A67'] = 'WINTER PIVOT SEASONALITY (property care - gutter, leaf, patio)'
ws_in['A67'].font = SECTION_FONT
ws_in['A68'] = 'Used by Sens - Winter Pivot scenario. Oct–Nov boosted for leaf/gutter; winter churn reduced (customers stay on year-round plan).'
ws_in['A68'].font = NOTE_FONT

for r in range(55, 67):
    mo = int(ws_in[f'A{r}'].value)
    base_acq = ws_in[f'B{r}'].value
    base_ch = ws_in[f'D{r}'].value
    p_acq, p_ch = WINTER_PIVOT[mo]
    ws_in[f'E{r}'] = p_acq
    ws_in[f'F{r}'] = p_ch
    ws_in[f'E{r}'].fill = INPUT_FILL
    ws_in[f'F{r}'].fill = INPUT_FILL
    note = ws_in[f'C{r}'].value or ''
    if p_acq != base_acq or p_ch != base_ch:
        ws_in[f'C{r}'] = f"{note.split('|')[0].strip()} | pivot acq ×{p_acq/base_acq:.2f}, churn ×{p_ch/base_ch:.2f}"

for name, ref in [
    ('WinterPivotSeasonTable', 'Inputs!$A$55:$E$66'),
    ('WinterPivotChurnTable', 'Inputs!$A$55:$F$66'),
]:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=ref))

# --- Scenario engine helpers (from rebuild, with season overrides) ---
def cl(sc, n):
    return get_column_letter(sc + COL[n] - 1)

def ch_expr(sc, r, ach):
    C = lambda n: cl(sc, n)
    return f"({ach}*VLOOKUP({C('CalMo')}{r},WinterPivotChurnTable,6,FALSE))"

def monthly_active_scalar(sc, r, new, act):
    C = lambda n: cl(sc, n)
    m, a, f = C(new), C(act), C('EffChurn')
    if r == FIRST:
        return f"={m}{r}"
    if r == FIRST + 1:
        return f"={m}{r}+{m}{r-1}"
    if r == FIRST + 2:
        return f"={m}{r}+{m}{r-1}+{m}{r-2}"
    return f"={m}{r}+{m}{r-1}+{m}{r-2}+MAX(0,{a}{r-1}-{m}{r-1}-{m}{r-2})*(1-{f}{r})"

def anniversary_terms(sc, r, new, ach, template):
    C = lambda n: cl(sc, n)
    o = C(new)
    ch = ch_expr(sc, r, ach)
    parts = []
    for yrs, offset in [(0, 12), (1, 24), (2, 36), (3, 48)]:
        src = r - offset
        if src < FIRST:
            continue
        parts.append(template(o, src, ch, yrs))
    return parts

def annual_churn_sum(sc, r, new, ach):
    parts = anniversary_terms(sc, r, new, ach, lambda o, src, ch, yrs: f"{o}{src}*{ch}*POWER(1-{ch},{yrs})")
    return "=0" if not parts else "=" + "+".join(parts)

def annual_active_scalar(sc, r, new, act, ach):
    C = lambda n: cl(sc, n)
    o, a = C(new), C(act)
    if r == FIRST:
        return f"={o}{r}"
    if r <= FIRST + 11:
        return f"=SUM({o}$5:{o}{r})"
    churn = annual_churn_sum(sc, r, new, ach).lstrip('=')
    return f"={a}{r-1}+{o}{r}-{churn}"

def ren_ann_scalar(sc, r, ach):
    C = lambda n: cl(sc, n)
    inf = C('Infl')
    def part(new, price):
        parts = anniversary_terms(sc, r, new, ach, lambda o, src, ch, yrs: f"{o}{src}*POWER(1-{ch},{yrs+1})*{price}*{inf}{r}")
        return "+".join(parts) if parts else "0"
    return f"={part('NEsA', SHARED['pessa'])}+{part('NPrA', SHARED['pprma'])}"

def vat_pay(sc, r):
    B, C, AD, AB = cl(sc, 'CalMo'), cl(sc, 'CalYear'), cl(sc, 'VATAccr'), cl(sc, 'VATReg')
    cy, cm, va = f"${C}${FIRST}:${C}{r}", f"${B}${FIRST}:${B}{r}", f"${AD}${FIRST}:${AD}{r}"
    cr = f"${C}{r}"
    q1 = f"SUMPRODUCT(--(({cy}={cr}-1)*({cm}>=10)*({cm}<=12)),{va})"
    q2 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=1)*({cm}<=3)),{va})"
    q3 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=4)*({cm}<=6)),{va})"
    q4 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=7)*({cm}<=9)),{va})"
    return f"=IF({AB}{r}=0,0,IF({B}{r}=1,{q1},IF({B}{r}=4,{q2},IF({B}{r}=7,{q3},IF({B}{r}=10,{q4},0)))))"

def write_winter_block(ws, sc, label, essmix, prem_uplift, hdr_row=3):
    ws.cell(2, sc, label).font = Font(bold=True, size=9, color="2F5496")
    hdrs = ["Mo","Cal","CY","Acq","Infl","ECh","Mkt","Paid","Ref","Tot","NMo","NAn",
            "NEsM","NPrM","NEsA","NPrA","AEsM","APrM","AEsA","APrA","Cust","MRR","ARR",
            "Mo$","NwAn","Ren","Gross","R12","VAT","Net","VAcc","VPay","VBal","Prov",
            "Fee","Ops","Fix","Ref$","Mkt","Pft","Cash","ARPU","Chrn"]
    for i, h in enumerate(hdrs):
        c = ws.cell(hdr_row, sc + i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    em = essmix
    pu = prem_uplift
    prem_m = f"({SHARED['pprmm']}+{pu})"
    prem_a = f"({SHARED['pprma']}+{pu}*12)"

    for m in range(1, 61):
        r = FIRST + m - 1
        C = lambda n: cl(sc, n)
        ws.cell(r, sc + COL['Mo'] - 1, m)
        ws.cell(r, sc + COL['CalMo'] - 1, f"=MOD({C('Mo')}{r}-1+LaunchMonth-1,12)+1")
        ws.cell(r, sc + COL['CalYear'] - 1, f"=INT(({C('Mo')}{r}-1+LaunchMonth-1)/12)+1")
        ws.cell(r, sc + COL['AcqSeas'] - 1, f"=VLOOKUP({C('CalMo')}{r},WinterPivotSeasonTable,5,FALSE)")
        ws.cell(r, sc + COL['Infl'] - 1, f"=(1+InflationRate)^INT(({C('Mo')}{r}-1)/12)")
        ws.cell(r, sc + COL['EffChurn'] - 1, f"=MonthlyChurn*VLOOKUP({C('CalMo')}{r},WinterPivotChurnTable,6,FALSE)")

        cb = f"{C('Cash')}{r-1}" if m > 1 else SHARED['init']
        rp = f"{C('GrossRev')}{r-1}" if m > 1 else "0"
        if m <= 3:
            mkt = f"={SHARED['mkt13']}"
        elif m <= 6:
            mkt = f"=MIN({SHARED['mkt46']},MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 12:
            mkt = f"=MIN(MAX({SHARED['min712']},{rp}*{SHARED['pct712']}),MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 18:
            mkt = f"=MIN(MAX({SHARED['min1318']},{rp}*{SHARED['pct1318']}),MAX(0,{cb}-{SHARED['buffer']}))"
        else:
            mkt = f"=MIN(MAX({SHARED['min19']},{rp}*{SHARED['pct19']}),MAX(0,{cb}-{SHARED['buffer']}))"
        ws.cell(r, sc + COL['Mkt'] - 1, mkt)
        ws.cell(r, sc + COL['PaidNew'] - 1, f"=IF({C('Mkt')}{r}>0,MAX(1,ROUND({C('Mkt')}{r}/CAC*{C('AcqSeas')}{r},0)),0)")
        ws.cell(r, sc + COL['RefNew'] - 1, f"=IF({C('PaidNew')}{r}>0,ROUND({C('PaidNew')}{r}*{SHARED['refmix']}/(1-{SHARED['refmix']}),0),0)")
        ws.cell(r, sc + COL['TotNew'] - 1, f"={C('PaidNew')}{r}+{C('RefNew')}{r}")
        ws.cell(r, sc + COL['NewMo'] - 1, f"=ROUND({C('TotNew')}{r}*(1-AnnualMix),0)")
        ws.cell(r, sc + COL['NewAnn'] - 1, f"={C('TotNew')}{r}-{C('NewMo')}{r}")
        ws.cell(r, sc + COL['NEsM'] - 1, f"=ROUND({C('NewMo')}{r}*{em},0)")
        ws.cell(r, sc + COL['NPrM'] - 1, f"={C('NewMo')}{r}-{C('NEsM')}{r}")
        ws.cell(r, sc + COL['NEsA'] - 1, f"=ROUND({C('NewAnn')}{r}*{em},0)")
        ws.cell(r, sc + COL['NPrA'] - 1, f"={C('NewAnn')}{r}-{C('NEsA')}{r}")
        ws.cell(r, sc + COL['AEsM'] - 1, monthly_active_scalar(sc, r, 'NEsM', 'AEsM'))
        ws.cell(r, sc + COL['APrM'] - 1, monthly_active_scalar(sc, r, 'NPrM', 'APrM'))
        ws.cell(r, sc + COL['AEsA'] - 1, annual_active_scalar(sc, r, 'NEsA', 'AEsA', 'AnnualChurn'))
        ws.cell(r, sc + COL['APrA'] - 1, annual_active_scalar(sc, r, 'NPrA', 'APrA', 'AnnualChurn'))

        inf = f"{C('Infl')}{r}"
        ws.cell(r, sc + COL['TotCust'] - 1, f"={C('AEsM')}{r}+{C('APrM')}{r}+{C('AEsA')}{r}+{C('APrA')}{r}")
        ws.cell(r, sc + COL['MRR'] - 1, f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{prem_m}*{inf}+{C('AEsA')}{r}*{SHARED['pessa']}*{inf}/12+{C('APrA')}{r}*{prem_a}*{inf}/12")
        ws.cell(r, sc + COL['ARR'] - 1, f"={C('MRR')}{r}*12")
        ws.cell(r, sc + COL['MoBill'] - 1, f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{prem_m}*{inf}")
        ws.cell(r, sc + COL['NewAnnCash'] - 1, f"={C('NEsA')}{r}*{SHARED['pessa']}*{inf}+{C('NPrA')}{r}*{prem_a}*{inf}")
        ws.cell(r, sc + COL['RenAnn'] - 1, ren_ann_scalar(sc, r, 'AnnualChurn'))
        ws.cell(r, sc + COL['GrossRev'] - 1, f"={C('MoBill')}{r}+{C('NewAnnCash')}{r}+{C('RenAnn')}{r}")
        ws.cell(r, sc + COL['Roll12'] - 1, f"=SUM({C('GrossRev')}${FIRST}:{C('GrossRev')}{r})" if m < 12 else f"=SUM({C('GrossRev')}{r-11}:{C('GrossRev')}{r})")
        ws.cell(r, sc + COL['VATReg'] - 1, f"={C('Roll12')}{r}>=VATThreshold" if m == 1 else f"=OR({C('VATReg')}{r-1},{C('Roll12')}{r}>=VATThreshold)")
        ws.cell(r, sc + COL['NetRev'] - 1, f"=IF({C('VATReg')}{r},{C('GrossRev')}{r}/(1+VATRate),{C('GrossRev')}{r})")
        ws.cell(r, sc + COL['VATAccr'] - 1, f"={C('GrossRev')}{r}-{C('NetRev')}{r}")
        ws.cell(r, sc + COL['VATPay'] - 1, vat_pay(sc, r))
        ws.cell(r, sc + COL['VATBal'] - 1, f"={C('VATAccr')}{r}-{C('VATPay')}{r}" if m == 1 else f"={C('VATBal')}{r-1}+{C('VATAccr')}{r}-{C('VATPay')}{r}")

        if m == 1:
            prov = f"=IF({SHARED['pdelay']}>=1,0,{C('NetRev')}{r}*{SHARED['prov']})"
        else:
            prov = f"=IF({SHARED['pdelay']}>=1,{C('NetRev')}{r-1}*{SHARED['prov']},{C('NetRev')}{r}*{SHARED['prov']})"
        ws.cell(r, sc + COL['Prov'] - 1, prov)
        ws.cell(r, sc + COL['PayFee'] - 1, f"={C('GrossRev')}{r}*{SHARED['pfee']}")
        ws.cell(r, sc + COL['OpsAI'] - 1, f"={C('TotCust')}{r}*{SHARED['ops']}")
        ws.cell(r, sc + COL['FixedCost'] - 1, f"={SHARED['fixed']}")
        ws.cell(r, sc + COL['RefCost'] - 1, f"={C('RefNew')}{r}*{SHARED['refreward']}")
        ws.cell(r, sc + COL['MktOut'] - 1, f"={C('Mkt')}{r}")
        ws.cell(r, sc + COL['Profit'] - 1, f"={C('NetRev')}{r}-{C('Prov')}{r}-{C('PayFee')}{r}-{C('OpsAI')}{r}-{C('FixedCost')}{r}-{C('RefCost')}{r}-{C('MktOut')}{r}")
        out = f"{C('Prov')}{r}+{C('PayFee')}{r}+{C('OpsAI')}{r}+{C('FixedCost')}{r}+{C('RefCost')}{r}+{C('MktOut')}{r}+{C('VATPay')}{r}"
        ws.cell(r, sc + COL['Cash'] - 1, f"={SHARED['init']}+{C('GrossRev')}{r}-({out})" if m == 1 else f"={C('Cash')}{r-1}+{C('GrossRev')}{r}-({out})")
        ws.cell(r, sc + COL['ARPU'] - 1, f"=IF({C('TotCust')}{r}>0,{C('MRR')}{r}/{C('TotCust')}{r},0)")
        ws.cell(r, sc + COL['Churn'] - 1, 0 if m == 1 else f"=IF({C('TotCust')}{r-1}>0,MAX(0,({C('TotCust')}{r-1}+{C('TotNew')}{r}-{C('TotCust')}{r})/{C('TotCust')}{r-1}),0)")

        for ci in range(sc, sc + len(NAMES)):
            ws.cell(r, ci).border = BORDER
            ncol = ci - sc + 1
            if ncol in (COL['MRR'], COL['ARR'], COL['GrossRev'], COL['NetRev'], COL['Profit'], COL['Cash']):
                ws.cell(r, ci).number_format = MONEY
            elif ncol in (COL['Churn'], COL['EffChurn']):
                ws.cell(r, ci).number_format = PCT2

    return {k: cl(sc, k) for k in ['TotCust', 'MRR', 'ARR', 'GrossRev', 'Profit', 'Cash']}

# --- Scenario Engine: add block 27 (after 26 existing blocks) ---
ws_se = wb['Scenario Engine']
WINTER_SC = 1 + BLOCK_W * 26  # col 1145
# Clear any prior winter block columns
for r in range(1, 65):
    for c in range(WINTER_SC, WINTER_SC + BLOCK_W):
        ws_se.cell(r, c).value = None

WP = write_winter_block(
    ws_se, WINTER_SC, 'Winter Pivot',
    essmix="'Sens - Winter Pivot'!$B$6",
    prem_uplift="'Sens - Winter Pivot'!$B$7",
)

# Base block refs (block index 1, sc=45)
BASE_SC = 1 + BLOCK_W
BASE = {k: cl(BASE_SC, k) for k in ['TotCust', 'MRR', 'ARR', 'GrossRev', 'Profit', 'Cash']}

# --- Sens - Winter Pivot sheet ---
if 'Sens - Winter Pivot' in wb.sheetnames:
    del wb['Sens - Winter Pivot']
idx = wb.sheetnames.index('Sens - CAC & Marketing') + 1
ws = wb.create_sheet('Sens - Winter Pivot', idx)

ws['A1'] = 'Winter Pivot - Property Care Scenario (gutter, leaf, patio)'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Compares base garden seasonality vs year-round property care. Edit yellow cells. Pivot seasonality: Inputs cols E & F.'
ws['A2'].font = NOTE_FONT

ws['A4'] = 'SCENARIO ASSUMPTIONS'
ws['A4'].font = SECTION_FONT
labels = [
    ('A5', 'Positioning', 'B5', 'Year-round property care subscription (mowing pauses Dec–Feb; gutter, leaf & patio included on Premium)'),
    ('A6', 'Essential tier mix', 'B6', 0.65),
    ('A7', 'Premium price uplift (£/mo)', 'B7', 10),
    ('A8', 'Same as base', 'B8', 'CAC, churn rates, marketing & pricing otherwise from Inputs'),
]
for a, la, b, val in labels:
    ws[a] = la
    ws[b] = val
    if isinstance(val, (int, float)):
        ws[b].fill = INPUT_FILL
        if b == 'B6':
            ws[b].number_format = '0%'
        if b == 'B7':
            ws[b].number_format = '£#,##0'

ws['A11'] = 'SEASONALITY COMPARISON (calendar month)'
ws['A11'].font = SECTION_FONT
headers = ['Month', 'Base acq', 'Pivot acq', 'Δ acq', 'Base churn×', 'Pivot churn×', 'Δ churn']
for i, h in enumerate(headers, 1):
    c = ws.cell(12, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

for i, mo in enumerate(range(1, 13)):
    r = 13 + i
    ws.cell(r, 1, MONTH_NAMES[mo - 1])
    ir = 54 + mo
    ws.cell(r, 2, f"=Inputs!B{ir}")
    ws.cell(r, 3, f"=Inputs!E{ir}")
    ws.cell(r, 4, f"=C{r}-B{r}")
    ws.cell(r, 5, f"=Inputs!D{ir}")
    ws.cell(r, 6, f"=Inputs!F{ir}")
    ws.cell(r, 7, f"=F{r}-E{r}")
    for c in range(1, 8):
        ws.cell(r, c).border = BORDER
        if c in (2, 3, 5, 6):
            ws.cell(r, c).number_format = '0.00'

ws['A27'] = 'OUTPUT COMPARISON - Base vs Winter Pivot (full model engine)'
ws['A27'].font = SECTION_FONT
ws['A28'] = 'Base = standard seasonality (Inputs cols B & D). Winter Pivot = cols E & F + Premium property care bundle.'
ws['A28'].font = NOTE_FONT

out_hdr = ['Metric', 'Base', 'Winter Pivot', 'Delta']
for i, h in enumerate(out_hdr, 1):
    c = ws.cell(30, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL

metrics = [
    ('End Year 1 - Customers', f"='Scenario Engine'!{BASE['TotCust']}$16", f"='Scenario Engine'!{WP['TotCust']}$16"),
    ('End Year 1 - MRR (£)', f"='Scenario Engine'!{BASE['MRR']}$16", f"='Scenario Engine'!{WP['MRR']}$16"),
    ('End Year 1 - ARR (£)', f"='Scenario Engine'!{BASE['ARR']}$16", f"='Scenario Engine'!{WP['ARR']}$16"),
    ('End Year 1 - Cash (£)', f"='Scenario Engine'!{BASE['Cash']}$16", f"='Scenario Engine'!{WP['Cash']}$16"),
    ('Year 1 - Gross Revenue (£)', f"=SUM('Scenario Engine'!{BASE['GrossRev']}$5:'Scenario Engine'!{BASE['GrossRev']}$16)", f"=SUM('Scenario Engine'!{WP['GrossRev']}$5:'Scenario Engine'!{WP['GrossRev']}$16)"),
    ('Year 1 - Profit (£)', f"=SUM('Scenario Engine'!{BASE['Profit']}$5:'Scenario Engine'!{BASE['Profit']}$16)", f"=SUM('Scenario Engine'!{WP['Profit']}$5:'Scenario Engine'!{WP['Profit']}$16)"),
    ('End Year 2 - Customers', f"='Scenario Engine'!{BASE['TotCust']}$28", f"='Scenario Engine'!{WP['TotCust']}$28"),
    ('End Year 2 - MRR (£)', f"='Scenario Engine'!{BASE['MRR']}$28", f"='Scenario Engine'!{WP['MRR']}$28"),
    ('End Year 2 - Cash (£)', f"='Scenario Engine'!{BASE['Cash']}$28", f"='Scenario Engine'!{WP['Cash']}$28"),
    ('End Year 5 - Customers', f"='Scenario Engine'!{BASE['TotCust']}$64", f"='Scenario Engine'!{WP['TotCust']}$64"),
    ('End Year 5 - MRR (£)', f"='Scenario Engine'!{BASE['MRR']}$64", f"='Scenario Engine'!{WP['MRR']}$64"),
    ('End Year 5 - Cash (£)', f"='Scenario Engine'!{BASE['Cash']}$64", f"='Scenario Engine'!{WP['Cash']}$64"),
]
for i, (label, base_f, wp_f) in enumerate(metrics):
    r = 31 + i
    ws.cell(r, 1, label)
    ws.cell(r, 2, base_f)
    ws.cell(r, 3, wp_f)
    ws.cell(r, 4, f"=C{r}-B{r}")
    for c in (2, 3, 4):
        ws.cell(r, c).number_format = MONEY if '£' in label or 'Revenue' in label or 'Profit' in label or 'Cash' in label else '#,##0'
    if 'Customers' in label:
        for c in (2, 3, 4):
            ws.cell(r, c).number_format = '#,##0'

ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 12

# Assumptions note
if 'Assumptions' in wb.sheetnames:
    ws_a = wb['Assumptions']
    ws_a['A29'] = 'Winter pivot seasonality'
    ws_a['B29'] = 'Inputs cols E & F'
    ws_a['C29'] = 'Sens - Winter Pivot tab - gutter/leaf/patio Oct–Feb'

# Headlines checklist
if 'Headlines' in wb.sheetnames:
    ws_h = wb['Headlines']
    for r in range(30, 40):
        if ws_h[f'A{r}'].value and 'Sensitivity' in str(ws_h[f'A{r}'].value):
            ws_h[f'A{r}'] = '✓ Sensitivity tabs - churn, mix, CAC, scenarios, winter pivot'
            break

wb.save(OUT)
print(f"Winter Pivot block at col {WINTER_SC} ({get_column_letter(WINTER_SC)})")
print(f"Base refs: {BASE}")
print(f"Winter refs: {WP}")
print("Saved Sens - Winter Pivot + Inputs cols E/F")
