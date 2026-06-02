#!/usr/bin/env python3
"""Build SaaS forecast v3/v4: Elite tier, annual visit cadence (10/20/30), VAT principal vs agent."""
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
PLANNING = ROOT / "planning"
SRC = PLANNING / "sorted_saas_recurring_revenue_forecast.xlsx"
OUT = PLANNING / "sorted_saas_recurring_revenue_forecast_v3_elite.xlsx"

# Annual visits per plan (product truth 2026-06) - provider pay = visits/12 × per-visit rate
VISITS_ESS_YR = 10
VISITS_PREM_YR = 20
VISITS_ELITE_YR = 30
PROV_SMALL = 20
PROV_MEDIUM = 30
PROV_LARGE = 40
ESS_SMALL_MO = 59.99
PREM_SMALL_MO = 84.99
ELITE_SMALL_MO = 119.99
GARDEN_MEDIUM_UPLIFT_MO = 20.0
GARDEN_LARGE_UPLIFT_MO = 40.0

FIRST = 5
BLOCK_W = 48
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, color="666666", size=10)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = "£#,##0"
PCT2 = "0.00%"

NAMES = [
    "Mo", "CalMo", "CalYear", "AcqSeas", "Infl", "EffChurn", "Mkt", "PaidNew", "RefNew", "TotNew",
    "NewMo", "NewAnn", "NEsM", "NPrM", "NElM", "NEsA", "NPrA", "NElA",
    "AEsM", "APrM", "AElM", "AEsA", "APrA", "AElA",
    "TotCust", "MRR", "ARR", "MoBill", "NewAnnCash", "RenAnn", "GrossRev", "Roll12", "VATReg", "NetRev",
    "VATAccr", "VATPay", "VATBal", "Prov", "PayFee", "OpsAI", "FixedCost", "RefCost", "MktOut",
    "Profit", "Cash", "ARPU", "Churn",
]
COL = {n: i for i, n in enumerate(NAMES, 1)}

SHARED = {
    "pessm": "BlendedPriceEssMonthly",
    "pprmm": "BlendedPricePremMonthly",
    "pelmm": "BlendedPriceEliteMonthly",
    "pessa": "BlendedPriceEssAnnual",
    "pprma": "BlendedPricePremAnnual",
    "pela": "BlendedPriceEliteAnnual",
    "paye": "BlendedProvPayEss",
    "payp": "BlendedProvPayPrem",
    "payel": "BlendedProvPayElite",
    "pfee": "PaymentFee",
    "ops": "OpsAI",
    "fixed": "MonthlyFixed",
    "init": "InitialInvestment",
    "buffer": "CashBuffer",
    "refmix": "ReferralMix",
    "refreward": "ReferralReward",
    "mkt13": "MktFixed1_3",
    "mkt46": "MktFixed4_6",
    "pct712": "MktPct7_12",
    "pct1318": "MktPct13_18",
    "pct19": "MktPct19plus",
    "min712": "MktMin7_12",
    "min1318": "MktMin13_18",
    "min19": "MktMin19plus",
    "essmix": "EssentialMix",
    "premmix": "PremiumMix",
    "elitemix": "EliteMix",
    "pdelay": "ProviderDelay",
}

METRICS = ["TotCust", "MRR", "ARR", "GrossRev", "NetRev", "Profit", "Cash", "PaidNew", "RefNew", "VATPay"]


def cl(sc, n):
    return get_column_letter(sc + COL[n] - 1)


def block_sc(index):
    return 1 + BLOCK_W * index


def se_ref(block_index, metric, row):
    return f"'Scenario Engine'!{cl(block_sc(block_index), metric)}${row}"


def se_sum(block_index, metric, r1, r2):
    c = cl(block_sc(block_index), metric)
    return f"=SUM('Scenario Engine'!{c}${r1}:'Scenario Engine'!{c}${r2})"


def model_ref(metric, row):
    return f"=Model!{cl(1, metric)}{row}"


def unmerge_input_blockers(ws):
    for rng in ("A13:C13", "A17:C17", "A14:C14", "A20:C20", "A21:C21", "A27:C27"):
        try:
            ws.unmerge_cells(rng)
        except (KeyError, ValueError):
            pass


def ch_expr(sc, r, ach):
    C = lambda n: cl(sc, n)
    return f"({ach}*VLOOKUP({C('CalMo')}{r},ChurnSeasonTable,4,FALSE))"


def monthly_active_scalar(sc, r, new, act):
    C = lambda n: cl(sc, n)
    m, a, f = C(new), C(act), C("EffChurn")
    if r == FIRST:
        return f"={m}{r}"
    if r == FIRST + 1:
        return f"={m}{r}+{m}{r-1}"
    if r == FIRST + 2:
        return f"={m}{r}+{m}{r-1}+{m}{r-2}"
    return f"={m}{r}+{m}{r-1}+{m}{r-2}+MAX(0,{a}{r-1}-{m}{r-1}-{m}{r-2})*(1-{f}{r})"


def anniversary_terms(sc, r, new, ach, template):
    C = lambda n: cl(sc, n)
    o = C(new)
    ch = ch_expr(sc, r, ach)
    parts = []
    for yrs, offset in [(0, 12), (1, 24), (2, 36), (3, 48)]:
        src = r - offset
        if src < FIRST:
            continue
        parts.append(template(o, src, ch, yrs))
    return parts


def annual_churn_sum(sc, r, new, ach):
    parts = anniversary_terms(
        sc, r, new, ach,
        lambda o, src, ch, yrs: f"{o}{src}*{ch}*POWER(1-{ch},{yrs})",
    )
    return "=0" if not parts else "=" + "+".join(parts)


def annual_active_scalar(sc, r, new, act, ach):
    C = lambda n: cl(sc, n)
    o, a = C(new), C(act)
    if r == FIRST:
        return f"={o}{r}"
    if r <= FIRST + 11:
        return f"=SUM({o}$5:{o}{r})"
    churn = annual_churn_sum(sc, r, new, ach).lstrip("=")
    return f"={a}{r-1}+{o}{r}-{churn}"


def ren_ann_scalar(sc, r, ach):
    C = lambda n: cl(sc, n)
    inf = C("Infl")

    def part(new, price):
        parts = anniversary_terms(
            sc, r, new, ach,
            lambda o, src, ch, yrs: f"{o}{src}*POWER(1-{ch},{yrs+1})*{price}*{inf}{r}",
        )
        return "+".join(parts) if parts else "0"

    return (
        f"={part('NEsA', SHARED['pessa'])}+{part('NPrA', SHARED['pprma'])}"
        f"+{part('NElA', SHARED['pela'])}"
    )


def prov_expr(sc, r):
    C = lambda n: cl(sc, n)
    return (
        f"{C('AEsM')}{r}*{SHARED['paye']}+{C('APrM')}{r}*{SHARED['payp']}"
        f"+{C('AElM')}{r}*{SHARED['payel']}+{C('AEsA')}{r}*{SHARED['paye']}"
        f"+{C('APrA')}{r}*{SHARED['payp']}+{C('AElA')}{r}*{SHARED['payel']}"
    )


def roll12_gross(sc, r):
    C = cl(sc, "GrossRev")
    if r < FIRST + 11:
        return f"SUM(${C}${FIRST}:${C}{r})"
    return f"SUM(${C}{r-11}:${C}{r})"


def roll12_margin(sc, r):
    Cg, Cp = cl(sc, "GrossRev"), cl(sc, "Prov")
    if r < FIRST + 11:
        return f"SUM(${Cg}${FIRST}:${Cg}{r})-SUM(${Cp}${FIRST}:${Cp}{r})"
    return f"SUM(${Cg}{r-11}:${Cg}{r})-SUM(${Cp}{r-11}:${Cp}{r})"


def vat_reg_expr(sc, r):
    C = cl(sc, "VATReg")
    roll_g = roll12_gross(sc, r)
    roll_m = roll12_margin(sc, r)
    test = f"IF(VATAgentMode,{roll_m}>=VATThreshold,{roll_g}>=VATThreshold)"
    if r == FIRST:
        return f"={test}"
    return f"=OR({C}{r-1},{test})"


def net_rev_expr(sc, r):
    Cg, Cp, Cr = cl(sc, "GrossRev"), cl(sc, "Prov"), cl(sc, "VATReg")
    margin = f"({Cg}{r}-{Cp}{r})"
    return (
        f"=IF({Cr}{r},IF(VATAgentMode,{Cg}{r}-({margin}-{margin}/(1+VATRate)),"
        f"{Cg}{r}/(1+VATRate)),{Cg}{r})"
    )


def vat_pay(sc, r):
    B, C, AD, AB = cl(sc, "CalMo"), cl(sc, "CalYear"), cl(sc, "VATAccr"), cl(sc, "VATReg")
    cy, cm, va = f"${C}${FIRST}:${C}{r}", f"${B}${FIRST}:${B}{r}", f"${AD}${FIRST}:${AD}{r}"
    cr = f"${C}{r}"
    q1 = f"SUMPRODUCT(--(({cy}={cr}-1)*({cm}>=10)*({cm}<=12)),{va})"
    q2 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=1)*({cm}<=3)),{va})"
    q3 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=4)*({cm}<=6)),{va})"
    q4 = f"SUMPRODUCT(--(({cy}={cr})*({cm}>=7)*({cm}<=9)),{va})"
    return f"=IF({AB}{r}=0,0,IF({B}{r}=1,{q1},IF({B}{r}=4,{q2},IF({B}{r}=7,{q3},IF({B}{r}=10,{q4},0)))))"


def write_block(ws, sc, label, mchurn, achurn, annmix, cac, hdr_row=4, essmix_override=None):
    if label:
        ws.cell(2, sc, label).font = Font(bold=True, size=9, color="2F5496")
    hdrs = [
        "Mo", "Cal", "CY", "Acq", "Infl", "ECh", "Mkt", "Paid", "Ref", "Tot", "NMo", "NAn",
        "NEsM", "NPrM", "NElM", "NEsA", "NPrA", "NElA", "AEsM", "APrM", "AElM", "AEsA", "APrA", "AElA",
        "Cust", "MRR", "ARR", "Mo$", "NwAn", "Ren", "Gross", "R12", "VAT", "Net", "VAcc", "VPay", "VBal",
        "Prov", "Fee", "Ops", "Fix", "Ref$", "Mkt", "Pft", "Cash", "ARPU", "Chrn",
    ]
    for i, h in enumerate(hdrs):
        c = ws.cell(hdr_row, sc + i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    em = essmix_override or SHARED["essmix"]

    for m in range(1, 61):
        r = FIRST + m - 1
        C = lambda n: cl(sc, n)
        ws.cell(r, sc + COL["Mo"] - 1, m)
        ws.cell(r, sc + COL["CalMo"] - 1, f"=MOD({C('Mo')}{r}-1+LaunchMonth-1,12)+1")
        ws.cell(r, sc + COL["CalYear"] - 1, f"=INT(({C('Mo')}{r}-1+LaunchMonth-1)/12)+1")
        ws.cell(r, sc + COL["AcqSeas"] - 1, f"=VLOOKUP({C('CalMo')}{r},SeasonTable,2,FALSE)")
        ws.cell(r, sc + COL["Infl"] - 1, f"=(1+InflationRate)^INT(({C('Mo')}{r}-1)/12)")
        ws.cell(r, sc + COL["EffChurn"] - 1, f"={mchurn}*VLOOKUP({C('CalMo')}{r},ChurnSeasonTable,4,FALSE)")

        cb = f"{C('Cash')}{r-1}" if m > 1 else SHARED["init"]
        rp = f"{C('GrossRev')}{r-1}" if m > 1 else "0"
        if m <= 3:
            mkt = f"={SHARED['mkt13']}"
        elif m <= 6:
            mkt = f"=MIN({SHARED['mkt46']},MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 12:
            mkt = f"=MIN(MAX({SHARED['min712']},{rp}*{SHARED['pct712']}),MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 18:
            mkt = f"=MIN(MAX({SHARED['min1318']},{rp}*{SHARED['pct1318']}),MAX(0,{cb}-{SHARED['buffer']}))"
        else:
            mkt = f"=MIN(MAX({SHARED['min19']},{rp}*{SHARED['pct19']}),MAX(0,{cb}-{SHARED['buffer']}))"
        ws.cell(r, sc + COL["Mkt"] - 1, mkt)
        ws.cell(r, sc + COL["PaidNew"] - 1, f"=IF({C('Mkt')}{r}>0,MAX(1,ROUND({C('Mkt')}{r}/{cac}*{C('AcqSeas')}{r},0)),0)")
        ws.cell(r, sc + COL["RefNew"] - 1, f"=IF({C('PaidNew')}{r}>0,ROUND({C('PaidNew')}{r}*{SHARED['refmix']}/(1-{SHARED['refmix']}),0),0)")
        ws.cell(r, sc + COL["TotNew"] - 1, f"={C('PaidNew')}{r}+{C('RefNew')}{r}")
        ws.cell(r, sc + COL["NewMo"] - 1, f"=ROUND({C('TotNew')}{r}*(1-{annmix}),0)")
        ws.cell(r, sc + COL["NewAnn"] - 1, f"={C('TotNew')}{r}-{C('NewMo')}{r}")
        ws.cell(r, sc + COL["NEsM"] - 1, f"=ROUND({C('NewMo')}{r}*{em},0)")
        ws.cell(r, sc + COL["NPrM"] - 1, f"=ROUND({C('NewMo')}{r}*{SHARED['premmix']},0)")
        ws.cell(r, sc + COL["NElM"] - 1, f"={C('NewMo')}{r}-{C('NEsM')}{r}-{C('NPrM')}{r}")
        ws.cell(r, sc + COL["NEsA"] - 1, f"=ROUND({C('NewAnn')}{r}*{em},0)")
        ws.cell(r, sc + COL["NPrA"] - 1, f"=ROUND({C('NewAnn')}{r}*{SHARED['premmix']},0)")
        ws.cell(r, sc + COL["NElA"] - 1, f"={C('NewAnn')}{r}-{C('NEsA')}{r}-{C('NPrA')}{r}")
        ws.cell(r, sc + COL["AEsM"] - 1, monthly_active_scalar(sc, r, "NEsM", "AEsM"))
        ws.cell(r, sc + COL["APrM"] - 1, monthly_active_scalar(sc, r, "NPrM", "APrM"))
        ws.cell(r, sc + COL["AElM"] - 1, monthly_active_scalar(sc, r, "NElM", "AElM"))
        ws.cell(r, sc + COL["AEsA"] - 1, annual_active_scalar(sc, r, "NEsA", "AEsA", achurn))
        ws.cell(r, sc + COL["APrA"] - 1, annual_active_scalar(sc, r, "NPrA", "APrA", achurn))
        ws.cell(r, sc + COL["AElA"] - 1, annual_active_scalar(sc, r, "NElA", "AElA", achurn))

        inf = f"{C('Infl')}{r}"
        ws.cell(r, sc + COL["TotCust"] - 1, f"={C('AEsM')}{r}+{C('APrM')}{r}+{C('AElM')}{r}+{C('AEsA')}{r}+{C('APrA')}{r}+{C('AElA')}{r}")
        ws.cell(
            r, sc + COL["MRR"] - 1,
            f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{SHARED['pprmm']}*{inf}"
            f"+{C('AElM')}{r}*{SHARED['pelmm']}*{inf}+{C('AEsA')}{r}*{SHARED['pessa']}*{inf}/12"
            f"+{C('APrA')}{r}*{SHARED['pprma']}*{inf}/12+{C('AElA')}{r}*{SHARED['pela']}*{inf}/12",
        )
        ws.cell(r, sc + COL["ARR"] - 1, f"={C('MRR')}{r}*12")
        ws.cell(
            r, sc + COL["MoBill"] - 1,
            f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{SHARED['pprmm']}*{inf}+{C('AElM')}{r}*{SHARED['pelmm']}*{inf}",
        )
        ws.cell(
            r, sc + COL["NewAnnCash"] - 1,
            f"={C('NEsA')}{r}*{SHARED['pessa']}*{inf}+{C('NPrA')}{r}*{SHARED['pprma']}*{inf}+{C('NElA')}{r}*{SHARED['pela']}*{inf}",
        )
        ws.cell(r, sc + COL["RenAnn"] - 1, ren_ann_scalar(sc, r, achurn))
        ws.cell(r, sc + COL["GrossRev"] - 1, f"={C('MoBill')}{r}+{C('NewAnnCash')}{r}+{C('RenAnn')}{r}")
        ws.cell(
            r, sc + COL["Roll12"] - 1,
            f"=SUM({C('GrossRev')}${FIRST}:{C('GrossRev')}{r})" if m < 12 else f"=SUM({C('GrossRev')}{r-11}:{C('GrossRev')}{r})",
        )
        ws.cell(r, sc + COL["VATReg"] - 1, vat_reg_expr(sc, r))
        ws.cell(r, sc + COL["NetRev"] - 1, net_rev_expr(sc, r))
        ws.cell(r, sc + COL["VATAccr"] - 1, f"={C('GrossRev')}{r}-{C('NetRev')}{r}")
        ws.cell(r, sc + COL["VATPay"] - 1, vat_pay(sc, r))
        ws.cell(r, sc + COL["VATBal"] - 1, f"={C('VATAccr')}{r}-{C('VATPay')}{r}" if m == 1 else f"={C('VATBal')}{r-1}+{C('VATAccr')}{r}-{C('VATPay')}{r}")

        if m == 1:
            prov = f"=IF({SHARED['pdelay']}>=1,0,{prov_expr(sc, r)})"
        else:
            prov = f"=IF({SHARED['pdelay']}>=1,{prov_expr(sc, r-1)},{prov_expr(sc, r)})"
        ws.cell(r, sc + COL["Prov"] - 1, prov)
        ws.cell(r, sc + COL["PayFee"] - 1, f"={C('GrossRev')}{r}*{SHARED['pfee']}")
        ws.cell(r, sc + COL["OpsAI"] - 1, f"={C('TotCust')}{r}*{SHARED['ops']}")
        ws.cell(r, sc + COL["FixedCost"] - 1, f"={SHARED['fixed']}")
        ws.cell(r, sc + COL["RefCost"] - 1, f"={C('RefNew')}{r}*{SHARED['refreward']}")
        ws.cell(r, sc + COL["MktOut"] - 1, f"={C('Mkt')}{r}")
        ws.cell(
            r, sc + COL["Profit"] - 1,
            f"={C('NetRev')}{r}-{C('Prov')}{r}-{C('PayFee')}{r}-{C('OpsAI')}{r}-{C('FixedCost')}{r}-{C('RefCost')}{r}-{C('MktOut')}{r}",
        )
        out = f"{C('Prov')}{r}+{C('PayFee')}{r}+{C('OpsAI')}{r}+{C('FixedCost')}{r}+{C('RefCost')}{r}+{C('MktOut')}{r}+{C('VATPay')}{r}"
        ws.cell(r, sc + COL["Cash"] - 1, f"={SHARED['init']}+{C('GrossRev')}{r}-({out})" if m == 1 else f"={C('Cash')}{r-1}+{C('GrossRev')}{r}-({out})")
        ws.cell(r, sc + COL["ARPU"] - 1, f"=IF({C('TotCust')}{r}>0,{C('MRR')}{r}/{C('TotCust')}{r},0)")
        ws.cell(
            r, sc + COL["Churn"] - 1,
            0 if m == 1 else f"=IF({C('TotCust')}{r-1}>0,MAX(0,({C('TotCust')}{r-1}+{C('TotNew')}{r}-{C('TotCust')}{r})/{C('TotCust')}{r-1}),0)",
        )

        for ci in range(sc, sc + len(NAMES)):
            ws.cell(r, ci).border = BORDER
            ncol = ci - sc + 1
            if ncol in (COL["MRR"], COL["ARR"], COL["GrossRev"], COL["NetRev"], COL["Profit"], COL["Cash"]):
                ws.cell(r, ci).number_format = MONEY
            elif ncol in (COL["Churn"], COL["EffChurn"]):
                ws.cell(r, ci).number_format = PCT2

    return {k: cl(sc, k) for k in METRICS}


def write_winter_block(ws, sc, label, essmix, prem_uplift, hdr_row=3):
    ws.cell(2, sc, label).font = Font(bold=True, size=9, color="2F5496")
    hdrs = [
        "Mo", "Cal", "CY", "Acq", "Infl", "ECh", "Mkt", "Paid", "Ref", "Tot", "NMo", "NAn",
        "NEsM", "NPrM", "NElM", "NEsA", "NPrA", "NElA", "AEsM", "APrM", "AElM", "AEsA", "APrA", "AElA",
        "Cust", "MRR", "ARR", "Mo$", "NwAn", "Ren", "Gross", "R12", "VAT", "Net", "VAcc", "VPay", "VBal",
        "Prov", "Fee", "Ops", "Fix", "Ref$", "Mkt", "Pft", "Cash", "ARPU", "Chrn",
    ]
    for i, h in enumerate(hdrs):
        c = ws.cell(hdr_row, sc + i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    em = essmix
    prem_m = f"(BlendedPricePremMonthly+{prem_uplift})"
    prem_a = f"(BlendedPricePremAnnual+{prem_uplift}*12)"

    for m in range(1, 61):
        r = FIRST + m - 1
        C = lambda n: cl(sc, n)
        ws.cell(r, sc + COL["Mo"] - 1, m)
        ws.cell(r, sc + COL["CalMo"] - 1, f"=MOD({C('Mo')}{r}-1+LaunchMonth-1,12)+1")
        ws.cell(r, sc + COL["CalYear"] - 1, f"=INT(({C('Mo')}{r}-1+LaunchMonth-1)/12)+1")
        ws.cell(r, sc + COL["AcqSeas"] - 1, f"=VLOOKUP({C('CalMo')}{r},WinterPivotSeasonTable,5,FALSE)")
        ws.cell(r, sc + COL["Infl"] - 1, f"=(1+InflationRate)^INT(({C('Mo')}{r}-1)/12)")
        ws.cell(r, sc + COL["EffChurn"] - 1, f"=MonthlyChurn*VLOOKUP({C('CalMo')}{r},WinterPivotChurnTable,6,FALSE)")

        cb = f"{C('Cash')}{r-1}" if m > 1 else SHARED["init"]
        rp = f"{C('GrossRev')}{r-1}" if m > 1 else "0"
        if m <= 3:
            mkt = f"={SHARED['mkt13']}"
        elif m <= 6:
            mkt = f"=MIN({SHARED['mkt46']},MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 12:
            mkt = f"=MIN(MAX({SHARED['min712']},{rp}*{SHARED['pct712']}),MAX(0,{cb}-{SHARED['buffer']}))"
        elif m <= 18:
            mkt = f"=MIN(MAX({SHARED['min1318']},{rp}*{SHARED['pct1318']}),MAX(0,{cb}-{SHARED['buffer']}))"
        else:
            mkt = f"=MIN(MAX({SHARED['min19']},{rp}*{SHARED['pct19']}),MAX(0,{cb}-{SHARED['buffer']}))"
        ws.cell(r, sc + COL["Mkt"] - 1, mkt)
        ws.cell(r, sc + COL["PaidNew"] - 1, f"=IF({C('Mkt')}{r}>0,MAX(1,ROUND({C('Mkt')}{r}/CAC*{C('AcqSeas')}{r},0)),0)")
        ws.cell(r, sc + COL["RefNew"] - 1, f"=IF({C('PaidNew')}{r}>0,ROUND({C('PaidNew')}{r}*{SHARED['refmix']}/(1-{SHARED['refmix']}),0),0)")
        ws.cell(r, sc + COL["TotNew"] - 1, f"={C('PaidNew')}{r}+{C('RefNew')}{r}")
        ws.cell(r, sc + COL["NewMo"] - 1, f"=ROUND({C('TotNew')}{r}*(1-AnnualMix),0)")
        ws.cell(r, sc + COL["NewAnn"] - 1, f"={C('TotNew')}{r}-{C('NewMo')}{r}")
        ws.cell(r, sc + COL["NEsM"] - 1, f"=ROUND({C('NewMo')}{r}*{em},0)")
        ws.cell(r, sc + COL["NPrM"] - 1, f"=ROUND({C('NewMo')}{r}*{SHARED['premmix']},0)")
        ws.cell(r, sc + COL["NElM"] - 1, f"={C('NewMo')}{r}-{C('NEsM')}{r}-{C('NPrM')}{r}")
        ws.cell(r, sc + COL["NEsA"] - 1, f"=ROUND({C('NewAnn')}{r}*{em},0)")
        ws.cell(r, sc + COL["NPrA"] - 1, f"=ROUND({C('NewAnn')}{r}*{SHARED['premmix']},0)")
        ws.cell(r, sc + COL["NElA"] - 1, f"={C('NewAnn')}{r}-{C('NEsA')}{r}-{C('NPrA')}{r}")
        ws.cell(r, sc + COL["AEsM"] - 1, monthly_active_scalar(sc, r, "NEsM", "AEsM"))
        ws.cell(r, sc + COL["APrM"] - 1, monthly_active_scalar(sc, r, "NPrM", "APrM"))
        ws.cell(r, sc + COL["AElM"] - 1, monthly_active_scalar(sc, r, "NElM", "AElM"))
        ws.cell(r, sc + COL["AEsA"] - 1, annual_active_scalar(sc, r, "NEsA", "AEsA", "AnnualChurn"))
        ws.cell(r, sc + COL["APrA"] - 1, annual_active_scalar(sc, r, "NPrA", "APrA", "AnnualChurn"))
        ws.cell(r, sc + COL["AElA"] - 1, annual_active_scalar(sc, r, "NElA", "AElA", "AnnualChurn"))

        inf = f"{C('Infl')}{r}"
        ws.cell(r, sc + COL["TotCust"] - 1, f"={C('AEsM')}{r}+{C('APrM')}{r}+{C('AElM')}{r}+{C('AEsA')}{r}+{C('APrA')}{r}+{C('AElA')}{r}")
        ws.cell(
            r, sc + COL["MRR"] - 1,
            f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{prem_m}*{inf}+{C('AElM')}{r}*{SHARED['pelmm']}*{inf}"
            f"+{C('AEsA')}{r}*{SHARED['pessa']}*{inf}/12+{C('APrA')}{r}*{prem_a}*{inf}/12+{C('AElA')}{r}*{SHARED['pela']}*{inf}/12",
        )
        ws.cell(r, sc + COL["ARR"] - 1, f"={C('MRR')}{r}*12")
        ws.cell(r, sc + COL["MoBill"] - 1, f"={C('AEsM')}{r}*{SHARED['pessm']}*{inf}+{C('APrM')}{r}*{prem_m}*{inf}+{C('AElM')}{r}*{SHARED['pelmm']}*{inf}")
        ws.cell(r, sc + COL["NewAnnCash"] - 1, f"={C('NEsA')}{r}*{SHARED['pessa']}*{inf}+{C('NPrA')}{r}*{prem_a}*{inf}+{C('NElA')}{r}*{SHARED['pela']}*{inf}")
        ws.cell(r, sc + COL["RenAnn"] - 1, ren_ann_scalar(sc, r, "AnnualChurn"))
        ws.cell(r, sc + COL["GrossRev"] - 1, f"={C('MoBill')}{r}+{C('NewAnnCash')}{r}+{C('RenAnn')}{r}")
        ws.cell(
            r, sc + COL["Roll12"] - 1,
            f"=SUM({C('GrossRev')}${FIRST}:{C('GrossRev')}{r})" if m < 12 else f"=SUM({C('GrossRev')}{r-11}:{C('GrossRev')}{r})",
        )
        ws.cell(r, sc + COL["VATReg"] - 1, vat_reg_expr(sc, r))
        ws.cell(r, sc + COL["NetRev"] - 1, net_rev_expr(sc, r))
        ws.cell(r, sc + COL["VATAccr"] - 1, f"={C('GrossRev')}{r}-{C('NetRev')}{r}")
        ws.cell(r, sc + COL["VATPay"] - 1, vat_pay(sc, r))
        ws.cell(r, sc + COL["VATBal"] - 1, f"={C('VATAccr')}{r}-{C('VATPay')}{r}" if m == 1 else f"={C('VATBal')}{r-1}+{C('VATAccr')}{r}-{C('VATPay')}{r}")

        if m == 1:
            prov = f"=IF({SHARED['pdelay']}>=1,0,{prov_expr(sc, r)})"
        else:
            prov = f"=IF({SHARED['pdelay']}>=1,{prov_expr(sc, r-1)},{prov_expr(sc, r)})"
        ws.cell(r, sc + COL["Prov"] - 1, prov)
        ws.cell(r, sc + COL["PayFee"] - 1, f"={C('GrossRev')}{r}*{SHARED['pfee']}")
        ws.cell(r, sc + COL["OpsAI"] - 1, f"={C('TotCust')}{r}*{SHARED['ops']}")
        ws.cell(r, sc + COL["FixedCost"] - 1, f"={SHARED['fixed']}")
        ws.cell(r, sc + COL["RefCost"] - 1, f"={C('RefNew')}{r}*{SHARED['refreward']}")
        ws.cell(r, sc + COL["MktOut"] - 1, f"={C('Mkt')}{r}")
        ws.cell(
            r, sc + COL["Profit"] - 1,
            f"={C('NetRev')}{r}-{C('Prov')}{r}-{C('PayFee')}{r}-{C('OpsAI')}{r}-{C('FixedCost')}{r}-{C('RefCost')}{r}-{C('MktOut')}{r}",
        )
        out = f"{C('Prov')}{r}+{C('PayFee')}{r}+{C('OpsAI')}{r}+{C('FixedCost')}{r}+{C('RefCost')}{r}+{C('MktOut')}{r}+{C('VATPay')}{r}"
        ws.cell(r, sc + COL["Cash"] - 1, f"={SHARED['init']}+{C('GrossRev')}{r}-({out})" if m == 1 else f"={C('Cash')}{r-1}+{C('GrossRev')}{r}-({out})")
        ws.cell(r, sc + COL["ARPU"] - 1, f"=IF({C('TotCust')}{r}>0,{C('MRR')}{r}/{C('TotCust')}{r},0)")
        ws.cell(
            r, sc + COL["Churn"] - 1,
            0 if m == 1 else f"=IF({C('TotCust')}{r-1}>0,MAX(0,({C('TotCust')}{r-1}+{C('TotNew')}{r}-{C('TotCust')}{r})/{C('TotCust')}{r-1}),0)",
        )

        for ci in range(sc, sc + len(NAMES)):
            ws.cell(r, ci).border = BORDER
            ncol = ci - sc + 1
            if ncol in (COL["MRR"], COL["ARR"], COL["GrossRev"], COL["NetRev"], COL["Profit"], COL["Cash"]):
                ws.cell(r, ci).number_format = MONEY
            elif ncol in (COL["Churn"], COL["EffChurn"]):
                ws.cell(r, ci).number_format = PCT2

    return {k: cl(sc, k) for k in ["TotCust", "MRR", "ARR", "GrossRev", "Profit", "Cash"]}


def set_defined_name(wb, name, ref):
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def clear_all_defined_names(wb):
    for name in list(wb.defined_names.keys()):
        del wb.defined_names[name]


def update_inputs(wb):
    clear_all_defined_names(wb)
    ws = wb["Inputs"]
    ws["A1"] = "Sorted - SaaS Forecast v4 (10/20/30 visits per year + VAT principal/agent)"
    ws.insert_rows(12, 2)
    ws.insert_rows(18, 2)
    ws.insert_rows(20, 3)
    ws.insert_rows(32, 2)
    ws.insert_rows(39, 10)
    unmerge_input_blockers(ws)

    pricing = [
        (8, "Essential Monthly (£/mo, small ≤50 m²)", ESS_SMALL_MO, "10 visits/yr - medium +£20, large +£40"),
        (9, "Premium Monthly (£/mo, small ≤50 m²)", PREM_SMALL_MO, "20 visits/yr - medium +£20, large +£40"),
        (10, "Elite Monthly (£/mo, small ≤50 m²)", ELITE_SMALL_MO, "30 visits/yr - medium +£20, large +£40"),
        (11, "Essential Annual (£/yr, small)", round(ESS_SMALL_MO * 10, 2), "10× monthly (~2 months free)"),
        (12, "Premium Annual (£/yr, small)", round(PREM_SMALL_MO * 10, 2), "10× monthly (~2 months free)"),
        (13, "Elite Annual (£/yr, small)", round(ELITE_SMALL_MO * 10, 2), "10× monthly (~2 months free)"),
    ]
    for row, label, val, note in pricing:
        ws.cell(row, 1, label)
        ws.cell(row, 2, val).fill = INPUT_FILL
        ws.cell(row, 3, note)

    ws["A15"] = "PLAN MIX"
    ws["A15"].font = SECTION_FONT
    mix_rows = [
        (16, "Annual plan mix (%)", 0.2, "Industry B2C benchmark: 15–25%"),
        (17, "Essential tier mix (%)", 0.55, "Share of new customers within billing type"),
        (18, "Premium tier mix (%)", 0.32, "Share of new customers within billing type"),
        (19, "Elite tier mix (%)", 0.13, "Share of new customers within billing type"),
    ]
    for row, label, val, note in mix_rows:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, val)
        c.fill = INPUT_FILL
        c.number_format = "0%"
        ws.cell(row, 3, note)

    ws["A20"] = "GARDEN SIZE MIX"
    ws["A20"].font = SECTION_FONT
    garden_rows = [
        (21, "Small garden mix (%)", 0.5, "≤50 m² maintained area"),
        (22, "Medium garden mix (%)", 0.35, "≤100 m² (+£20/mo customer, +£10/visit provider vs small)"),
        (23, "Large garden mix (%)", 0.15, "≤150 m² (+£40/mo customer, +£20/visit provider vs small)"),
    ]
    for row, label, val, note in garden_rows:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, val)
        c.fill = INPUT_FILL
        c.number_format = "0%"
        ws.cell(row, 3, note)

    ws["A24"] = "CHURN (edit for sensitivity)"
    ws["A24"].font = SECTION_FONT
    ws["A25"] = "Monthly plan churn - post min term (%/mo)"
    ws["B25"] = 0.055
    ws["B25"].fill = INPUT_FILL
    ws["A26"] = "Annual plan churn - at renewal (%/yr)"
    ws["B26"] = 0.2
    ws["B26"].fill = INPUT_FILL

    ws["A28"] = "UNIT ECONOMICS"
    ws["A28"].font = SECTION_FONT
    ess_mo = round(VISITS_ESS_YR / 12 * PROV_SMALL, 2)
    prem_mo = round(VISITS_PREM_YR / 12 * PROV_SMALL, 2)
    elite_mo = round(VISITS_ELITE_YR / 12 * PROV_SMALL, 2)
    unit_rows = [
        (29, "Customer Acquisition Cost - CAC (£)", 70, "Paid acquisition per customer"),
        (30, "Provider pay - Essential (£/mo, small garden)", ess_mo, f"{VISITS_ESS_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (31, "Provider pay - Premium (£/mo, small garden)", prem_mo, f"{VISITS_PREM_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (32, "Provider pay - Elite (£/mo, small garden)", elite_mo, f"{VISITS_ELITE_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (33, "Payment processing fee (%)", 0.03, "Stripe/card fees"),
        (34, "Ops + AI cost per customer (£/mo)", 5, "Support, hosting, AI per active sub"),
    ]
    for row, label, val, note in unit_rows:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, val)
        c.fill = INPUT_FILL
        if "%" in label:
            c.number_format = "0%"
        ws.cell(row, 3, note)

    ws["A36"] = "CASH TIMING"
    ws["A36"].font = SECTION_FONT
    ws["A37"] = "Provider payment delay (months)"
    ws["B37"] = 1
    ws["B37"].fill = INPUT_FILL
    ws["A38"] = "Monthly fixed platform cost (£)"
    ws["B38"] = 250
    ws["B38"].fill = INPUT_FILL

    ws["A39"] = "BLENDED PRICING & PROVIDER PAY (calculated)"
    ws["A39"].font = SECTION_FONT
    ws["A40"] = "Blended Essential monthly (£/mo)"
    ws["B40"] = "=$B$8*$B$21+($B$8+10)*$B$22+($B$8+20)*$B$23"
    ws["A41"] = "Blended Premium monthly (£/mo)"
    ws["B41"] = "=$B$9*$B$21+($B$9+10)*$B$22+($B$9+20)*$B$23"
    ws["A42"] = "Blended Elite monthly (£/mo)"
    ws["B42"] = "=$B$10*$B$21+($B$10+10)*$B$22+($B$10+20)*$B$23"
    ws["A43"] = "Blended Essential annual (£/yr)"
    ws["B43"] = "=$B$11*$B$21+($B$11+100)*$B$22+($B$11+200)*$B$23"
    ws["A44"] = "Blended Premium annual (£/yr)"
    ws["B44"] = "=$B$12*$B$21+($B$12+100)*$B$22+($B$12+200)*$B$23"
    ws["A45"] = "Blended Elite annual (£/yr)"
    ws["B45"] = "=$B$13*$B$21+($B$13+100)*$B$22+($B$13+200)*$B$23"
    blend_visit = f"($B$21*{PROV_SMALL}+$B$22*{PROV_MEDIUM}+$B$23*{PROV_LARGE})"
    ws["A46"] = "Blended provider pay - Essential (£/mo)"
    ws["B46"] = f"=({VISITS_ESS_YR}/12)*{blend_visit}"
    ws["A47"] = "Blended provider pay - Premium (£/mo)"
    ws["B47"] = f"=({VISITS_PREM_YR}/12)*{blend_visit}"
    ws["A48"] = "Blended provider pay - Elite (£/mo)"
    ws["B48"] = f"=({VISITS_ELITE_YR}/12)*{blend_visit}"
    for row in range(40, 49):
        ws.cell(row, 2).number_format = MONEY

    # Season tables shifted vs original
    season_start = 74
    ws[f"A{season_start - 2}"] = "SEASONALITY (calendar month)"
    ws[f"A{season_start - 2}"].font = SECTION_FONT
    ws[f"A{season_start - 1}"] = "Acquisition multiplier (col B), notes (col C), churn multiplier (col D). Winter pivot cols E & F."
    ws[f"A{season_start - 1}"].font = NOTE_FONT
    headers = ["Month", "Acq mult", "Notes", "Churn mult", "Pivot acq", "Pivot churn"]
    for i, h in enumerate(headers, 1):
        ws.cell(season_start - 1, i, h).font = Font(bold=True)

    old_ws = openpyxl.load_workbook(SRC)["Inputs"]
    for mo in range(1, 13):
        old_r = 54 + mo
        new_r = season_start + mo - 1
        for col in range(1, 7):
            val = old_ws.cell(old_r, col).value
            if val is not None:
                ws.cell(new_r, col, val)

    names = {
        "InitialInvestment": "Inputs!$B$5",
        "PriceEssMonthly": "Inputs!$B$8",
        "PricePremMonthly": "Inputs!$B$9",
        "PriceEliteMonthly": "Inputs!$B$10",
        "PriceEssAnnual": "Inputs!$B$11",
        "PricePremAnnual": "Inputs!$B$12",
        "PriceEliteAnnual": "Inputs!$B$13",
        "AnnualMix": "Inputs!$B$16",
        "EssentialMix": "Inputs!$B$17",
        "PremiumMix": "Inputs!$B$18",
        "EliteMix": "Inputs!$B$19",
        "GardenSmallMix": "Inputs!$B$21",
        "GardenMediumMix": "Inputs!$B$22",
        "GardenLargeMix": "Inputs!$B$23",
        "MonthlyChurn": "Inputs!$B$25",
        "AnnualChurn": "Inputs!$B$26",
        "CAC": "Inputs!$B$29",
        "ProvPayEss": "Inputs!$B$30",
        "ProvPayPrem": "Inputs!$B$31",
        "ProvPayElite": "Inputs!$B$32",
        "PaymentFee": "Inputs!$B$33",
        "OpsAI": "Inputs!$B$34",
        "ProviderDelay": "Inputs!$B$37",
        "MonthlyFixed": "Inputs!$B$38",
        "BlendedPriceEssMonthly": "Inputs!$B$40",
        "BlendedPricePremMonthly": "Inputs!$B$41",
        "BlendedPriceEliteMonthly": "Inputs!$B$42",
        "BlendedPriceEssAnnual": "Inputs!$B$43",
        "BlendedPricePremAnnual": "Inputs!$B$44",
        "BlendedPriceEliteAnnual": "Inputs!$B$45",
        "BlendedProvPayEss": "Inputs!$B$46",
        "BlendedProvPayPrem": "Inputs!$B$47",
        "BlendedProvPayElite": "Inputs!$B$48",
        "MktFixed1_3": "Inputs!$B$51",
        "MktFixed4_6": "Inputs!$B$52",
        "MktPct7_12": "Inputs!$B$53",
        "MktPct13_18": "Inputs!$B$54",
        "MktPct19plus": "Inputs!$B$55",
        "MktMin7_12": "Inputs!$B$56",
        "MktMin13_18": "Inputs!$B$57",
        "MktMin19plus": "Inputs!$B$58",
        "CashBuffer": "Inputs!$B$59",
        "InflationRate": "Inputs!$B$60",
        "ReferralMix": "Inputs!$B$61",
        "ReferralReward": "Inputs!$B$62",
        "LaunchMonth": "Inputs!$B$63",
        "VATRate": "Inputs!$B$64",
        "VATThreshold": "Inputs!$B$65",
        "VATAgentMode": "Inputs!$B$66",
        "SeasonTable": f"Inputs!$A${season_start}:$B${season_start + 11}",
        "AcqSeasonTable": f"Inputs!$A${season_start}:$C${season_start + 11}",
        "ChurnSeasonTable": f"Inputs!$A${season_start}:$D${season_start + 11}",
        "WinterPivotSeasonTable": f"Inputs!$A${season_start}:$E${season_start + 11}",
        "WinterPivotChurnTable": f"Inputs!$A${season_start}:$F${season_start + 11}",
    }
    for name, ref in names.items():
        set_defined_name(wb, name, ref)

    # Copy marketing / VAT / inflation rows from old sheet with shifted refs
    shift_map = {
        31: 51, 32: 52, 33: 53, 34: 54, 35: 55, 36: 56, 37: 57, 38: 58, 39: 59,
        48: 60, 49: 61, 50: 62, 51: 63, 42: 64, 43: 65,
    }
    for old_r, new_r in shift_map.items():
        ws.cell(new_r, 1, old_ws.cell(old_r, 1).value)
        ws.cell(new_r, 2, old_ws.cell(old_r, 2).value)
        if old_ws.cell(old_r, 2).fill:
            ws.cell(new_r, 2).fill = copy(old_ws.cell(old_r, 2).fill)
        ws.cell(new_r, 3, old_ws.cell(old_r, 3).value)

    ws["A66"] = "VAT basis - Agent mode (1=Agent, 0=Principal)"
    ws["A66"].font = SECTION_FONT
    ws["B66"] = 0
    ws["B66"].fill = INPUT_FILL
    ws["C66"] = "0 = VAT on full customer revenue (principal). 1 = VAT on platform margin only (agent)."

    return season_start


def update_assumptions(wb):
    ws = wb["Assumptions"]
    ws["A1"] = "Assumptions & Methodology v4 (10/20/30 visits per year + VAT modes)"
    rows = [
        (6, "Essential Monthly", "=PriceEssMonthly", f"{VISITS_ESS_YR} visits/yr - 3-month minimum"),
        (7, "Premium Monthly", "=PricePremMonthly", f"{VISITS_PREM_YR} visits/yr - 3-month minimum"),
        (8, "Elite Monthly", "=PriceEliteMonthly", f"{VISITS_ELITE_YR} visits/yr + patio refresh - 3-month minimum"),
        (9, "Essential Annual", "=PriceEssAnnual", "12-month commitment"),
        (10, "Premium Annual", "=PricePremAnnual", "12-month commitment"),
        (11, "Elite Annual", "=PriceEliteAnnual", "12-month commitment"),
        (12, "Annual plan mix", "=AnnualMix", "Industry B2C: 15–25%"),
        (13, "Essential tier mix", "=EssentialMix", "Share of new customers"),
        (14, "Premium tier mix", "=PremiumMix", "Share of new customers"),
        (15, "Elite tier mix", "=EliteMix", "Share of new customers"),
        (16, "Small garden mix", "=GardenSmallMix", "Customer & provider blend"),
        (17, "Medium garden mix", "=GardenMediumMix", "+£10/mo customer, +£3/visit provider"),
        (18, "Large garden mix", "=GardenLargeMix", "+£20/mo customer, +£6/visit provider"),
        (19, "Monthly churn (post min term)", "=MonthlyChurn", "Home services B2C: 4–7%/mo"),
        (20, "Annual churn (at renewal)", "=AnnualChurn", "B2C annual: 15–25%"),
        (21, "CAC", "=CAC", "Paid acquisition"),
        (22, "Provider pay - Essential (small)", "=ProvPayEss", f"{VISITS_ESS_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (23, "Provider pay - Premium (small)", "=ProvPayPrem", f"{VISITS_PREM_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (24, "Provider pay - Elite (small)", "=ProvPayElite", f"{VISITS_ELITE_YR} visits/yr × £{PROV_SMALL}/visit ÷ 12"),
        (25, "Blended provider pay - Essential", "=BlendedProvPayEss", "Garden-size weighted"),
        (26, "Blended provider pay - Premium", "=BlendedProvPayPrem", "Garden-size weighted"),
        (27, "Blended provider pay - Elite", "=BlendedProvPayElite", "Garden-size weighted"),
        (28, "Payment fees", "=PaymentFee", "Stripe processing"),
        (29, "Ops + AI per customer", "=OpsAI", "Per active subscriber/month"),
        (30, "VAT agent mode", "=VATAgentMode", "0=Principal (VAT on gross), 1=Agent (VAT on margin)"),
    ]
    for row, param, val, note in rows:
        ws.cell(row, 1, param)
        ws.cell(row, 2, val)
        ws.cell(row, 3, note)


def write_vat_dual_calc(wb):
    if "VAT Dual Calc" in wb.sheetnames:
        del wb["VAT Dual Calc"]
    ws = wb.create_sheet("VAT Dual Calc")
    ws.sheet_state = "hidden"
    gcol, pcol = cl(1, "GrossRev"), cl(1, "Prov")
    hdrs = [
        "Mo", "Gross", "Prov", "Margin", "Roll12G", "Roll12M", "RegP", "RegA",
        "VatAccP", "VatAccA", "NetP", "NetA", "GP_P", "GP_A",
    ]
    for i, h in enumerate(hdrs, 1):
        ws.cell(4, i, h).font = Font(bold=True)

    for m in range(1, 61):
        r = FIRST + m - 1
        ws.cell(r, 1, m)
        ws.cell(r, 2, f"=Model!{gcol}{r}")
        ws.cell(r, 3, f"=Model!{pcol}{r}")
        ws.cell(r, 4, f"=B{r}-C{r}")
        if m < 12:
            ws.cell(r, 5, f"=SUM($B${FIRST}:B{r})")
            ws.cell(r, 6, f"=E{r}-SUM($C${FIRST}:C{r})")
        else:
            ws.cell(r, 5, f"=SUM(B{r-11}:B{r})")
            ws.cell(r, 6, f"=E{r}-SUM(C{r-11}:C{r})")
        if m == 1:
            ws.cell(r, 7, f"=E{r}>=VATThreshold")
            ws.cell(r, 8, f"=F{r}>=VATThreshold")
        else:
            ws.cell(r, 7, f"=OR(G{r-1},E{r}>=VATThreshold)")
            ws.cell(r, 8, f"=OR(H{r-1},F{r}>=VATThreshold)")
        ws.cell(r, 9, f"=IF(G{r},B{r}-B{r}/(1+VATRate),0)")
        ws.cell(r, 10, f"=IF(H{r},D{r}-D{r}/(1+VATRate),0)")
        ws.cell(r, 11, f"=B{r}-I{r}")
        ws.cell(r, 12, f"=B{r}-J{r}")
        ws.cell(r, 13, f"=K{r}-C{r}")
        ws.cell(r, 14, f"=L{r}-C{r}")
        for ci in range(1, 15):
            ws.cell(r, ci).number_format = MONEY if ci > 1 else "0"


def update_vat_comparison(wb):
    if "VAT Comparison" in wb.sheetnames:
        del wb["VAT Comparison"]
    idx = wb.sheetnames.index("Headlines") if "Headlines" in wb.sheetnames else 1
    ws = wb.create_sheet("VAT Comparison", idx + 1)
    ws["A1"] = "VAT - Principal vs Agent"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = (
        "Parallel comparison using Model gross revenue and provider pay. "
        "Toggle live model on Inputs B66: 0 = Principal (VAT on gross), 1 = Agent (VAT on margin only)."
    )
    ws["A3"].font = NOTE_FONT
    ws["A4"] = "See docs/vat-principal-vs-agent.md - accountant review required before filing."
    ws["A4"].font = NOTE_FONT
    hdrs = [
        "Year", "Gross revenue", "Provider pay", "Platform margin",
        "VAT registered (P)", "VAT registered (A)",
        "VAT accrued (P)", "VAT accrued (A)",
        "Net revenue (P)", "Net revenue (A)",
        "Gross profit (P)", "Gross profit (A)", "GP uplift (A vs P)",
    ]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(6, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    years = [(1, FIRST, 16), (2, 17, 28), (3, 29, 40), (4, 41, 52), (5, 53, 64)]
    for i, (year, start, end) in enumerate(years):
        r = 7 + i
        ws.cell(r, 1, f"Year {year}")
        ws.cell(r, 2, f"=SUM('VAT Dual Calc'!B{start}:B{end})")
        ws.cell(r, 3, f"=SUM('VAT Dual Calc'!C{start}:C{end})")
        ws.cell(r, 4, f"=B{r}-C{r}")
        ws.cell(r, 5, f"='VAT Dual Calc'!G{end}")
        ws.cell(r, 6, f"='VAT Dual Calc'!H{end}")
        ws.cell(r, 7, f"=SUM('VAT Dual Calc'!I{start}:I{end})")
        ws.cell(r, 8, f"=SUM('VAT Dual Calc'!J{start}:J{end})")
        ws.cell(r, 9, f"=SUM('VAT Dual Calc'!K{start}:K{end})")
        ws.cell(r, 10, f"=SUM('VAT Dual Calc'!L{start}:L{end})")
        ws.cell(r, 11, f"=SUM('VAT Dual Calc'!M{start}:M{end})")
        ws.cell(r, 12, f"=SUM('VAT Dual Calc'!N{start}:N{end})")
        ws.cell(r, 13, f"=L{r}-K{r}")
        for ci in range(2, 14):
            ws.cell(r, ci).number_format = MONEY if ci != 5 and ci != 6 else "0"

    ws["A14"] = "First VAT registration month (Principal)"
    ws["B14"] = f"=IFERROR(MATCH(TRUE,'VAT Dual Calc'!G{FIRST}:G{64},0)+{FIRST - 1},\"Not in 60 mo\")"
    ws["A15"] = "First VAT registration month (Agent)"
    ws["B15"] = f"=IFERROR(MATCH(TRUE,'VAT Dual Calc'!H{FIRST}:H{64},0)+{FIRST - 1},\"Not in 60 mo\")"
    ws["A17"] = "Current model mode"
    ws["B17"] = "=IF(VATAgentMode=1,\"Agent (margin VAT)\",\"Principal (gross VAT)\")"


def update_headlines(wb):
    ws = wb["Headlines"]
    ws["A1"] = "GardensSorted - Financial Headlines (v4 visits + VAT modes)"
    ws["A28"] = "✓ SaaS cohorts - Essential / Premium / Elite tiers, monthly & annual subs, MRR/ARR"
    ws["A29"] = f"✓ Visit cadence - {VISITS_ESS_YR}/{VISITS_PREM_YR}/{VISITS_ELITE_YR} visits per year (provider pay = visits/12 × per-visit rate)"
    ws["A30"] = "✓ Garden size mix - customer price & provider pay uplifts for medium (+£10/+£3) and large (+£20/+£6)"
    ws["A31"] = "✓ VAT - toggle Principal vs Agent on Inputs B66; see VAT Comparison sheet"
    ws["A32"] = None
    metrics = [
        ("A6", model_ref("ARR", 16)), ("C6", model_ref("ARR", 28)), ("E6", model_ref("ARR", 64)), ("G6", model_ref("Cash", 64)),
        ("B10", model_ref("TotCust", 16)), ("C10", model_ref("TotCust", 28)), ("D10", model_ref("TotCust", 64)),
        ("B11", model_ref("ARR", 16)), ("C11", model_ref("ARR", 28)), ("D11", model_ref("ARR", 64)),
        ("B12", "=SUM(Model!AE5:AE16)"), ("C12", "=SUM(Model!AE17:AE28)"), ("D12", "=SUM(Model!AE53:AE64)"),
        ("B13", "=SUM(Model!AH5:AH16)"), ("C13", "=SUM(Model!AH17:AH28)"), ("D13", "=SUM(Model!AH53:AH64)"),
        ("B14", "=SUM(Model!AJ5:AJ16)"), ("C14", "=SUM(Model!AJ17:AJ28)"), ("D14", "=SUM(Model!AJ53:AJ64)"),
        ("B15", "=SUM(Model!AR5:AR16)"), ("C15", "=SUM(Model!AR17:AR28)"), ("D15", "=SUM(Model!AR53:AR64)"),
        ("B16", model_ref("Cash", 16)), ("C16", model_ref("Cash", 28)), ("D16", model_ref("Cash", 64)),
        ("B17", "=AVERAGE(Model!AU6:AU16)"), ("C17", "=AVERAGE(Model!AU18:AU28)"), ("D17", "=AVERAGE(Model!AU54:AU64)"),
        ("B21", '=IFERROR("Month "&MATCH(TRUE,Model!AG5:Model!AG64,0),"Not in 60 mo")'),
        ("B22", '=IFERROR("Month "&MATCH(1,IF(Model!AR5:AR64>0,1),0),"Not in 60 mo")'),
        ("B23", '=IFERROR("Month "&MATCH(1,IF(Model!AS5:AS64>=InitialInvestment,1),0),"Not in 60 mo")'),
    ]
    for cell, formula in metrics:
        ws[cell] = formula


def update_five_year_summary(wb):
    ws = wb["5 Year Summary"]
    years = [
        (4, 16, 5, 16),
        (5, 28, 17, 28),
        (6, 40, 29, 40),
        (7, 52, 41, 52),
        (8, 64, 53, 64),
    ]
    for row, end, start, _ in years:
        ws[f"B{row}"] = model_ref("TotCust", end)
        ws[f"C{row}"] = model_ref("MRR", end)
        ws[f"D{row}"] = model_ref("ARR", end)
        ws[f"E{row}"] = f"=SUM(Model!H{start}:H{end})"
        ws[f"F{row}"] = f"=SUM(Model!I{start}:I{end})"
        ws[f"G{row}"] = f"=SUM(Model!AE{start}:AE{end})"
        ws[f"H{row}"] = f"=SUM(Model!AH{start}:AH{end})"
        ws[f"I{row}"] = f"=SUM(Model!AI{start}:AI{end})"
        ws[f"J{row}"] = f"=SUM(Model!AJ{start}:AJ{end})"
        ws[f"K{row}"] = model_ref("VATBal", end)
        ws[f"L{row}"] = f"=SUM(Model!AL{start}:AL{end})"
        ws[f"M{row}"] = f"=SUM(Model!AM{start}:AM{end})"
        ws[f"N{row}"] = f"=SUM(Model!AN{start}:AN{end})"
        ws[f"O{row}"] = f"=SUM(Model!AO{start}:AO{end})"
        ws[f"P{row}"] = f"=SUM(Model!AP{start}:AP{end})"
        ws[f"Q{row}"] = f"=SUM(Model!AQ{start}:AQ{end})"
        ws[f"R{row}"] = f"=SUM(Model!AR{start}:AR{end})"
        ws[f"S{row}"] = model_ref("Cash", end)
        ws[f"T{row}"] = f"=IF(G{row}>0,(H{row}-L{row}-M{row}-N{row}-O{row})/G{row},0)"
        ws[f"W{row}"] = f"=AVERAGE(Model!AU{start+1}:AU{end})"
    ws["B10"] = '=IFERROR("Month "&MATCH(TRUE,Model!AG5:Model!AG64,0),"Not in 60 mo")'


def update_24_month_forecast(wb):
    ws = wb["24 Month Forecast"]
    col_map = {
        6: "TotCust", 7: "EffChurn", 8: "MRR", 9: "ARR", 10: "GrossRev", 11: "NetRev",
        12: "VATAccr", 13: "VATPay", 14: "VATBal", 15: "Prov", 16: "PayFee", 17: "OpsAI",
        18: "FixedCost", 19: "RefCost", 20: "MktOut", 21: "Profit", 22: "Cash", 23: "Churn",
    }
    static = {2: "CalMo", 3: "PaidNew", 4: "RefNew", 5: "TotNew"}
    for m in range(1, 25):
        r = 3 + m
        model_row = 4 + m
        ws[f"A{r}"] = f"Month {m}"
        for c, metric in static.items():
            ws.cell(r, c, model_ref(metric, model_row))
        for c, metric in col_map.items():
            ws.cell(r, c, model_ref(metric, model_row))


def update_unit_economics(wb):
    ws = wb["Unit Economics"]
    ws["A1"] = "Unit Economics v4 (10/20/30 visits + garden-size blended)"
    ws["B4"] = (
        "=EssentialMix*(1-AnnualMix)*BlendedPriceEssMonthly+PremiumMix*(1-AnnualMix)*BlendedPricePremMonthly"
        "+EliteMix*(1-AnnualMix)*BlendedPriceEliteMonthly+EssentialMix*AnnualMix*BlendedPriceEssAnnual/12"
        "+PremiumMix*AnnualMix*BlendedPricePremAnnual/12+EliteMix*AnnualMix*BlendedPriceEliteAnnual/12"
    )
    ws["B5"] = "=EssentialMix*BlendedPriceEssMonthly+PremiumMix*BlendedPricePremMonthly+EliteMix*BlendedPriceEliteMonthly"
    ws["B6"] = "=EssentialMix*BlendedPriceEssAnnual/12+PremiumMix*BlendedPricePremAnnual/12+EliteMix*BlendedPriceEliteAnnual/12"
    ws["A8"] = "Provider cost per customer (blended tier + garden size)"
    ws["B8"] = "=EssentialMix*BlendedProvPayEss+PremiumMix*BlendedProvPayPrem+EliteMix*BlendedProvPayElite"
    ws["B12"] = "=B4-B8-B4*PaymentFee-OpsAI"
    ws["B18"] = "=B5*(1-PaymentFee)*ProvPayEss/PriceEssMonthly*(1/MonthlyChurn)*EssentialMix/(EssentialMix+PremiumMix+EliteMix)"
    ws["B19"] = "=B6*12*(1-PaymentFee)*(1/AnnualChurn)"
    ws["B21"] = "=B4*(1-PaymentFee)*((1-AnnualMix)/MonthlyChurn+AnnualMix*12/AnnualChurn)/CAC"


def update_sens_scenarios(wb):
    ws = wb["Sens - Scenarios"]
    ws["A1"] = "Scenario Comparison - Conservative / Base / Aggressive (v4)"
    blocks = {"B": 0, "C": 1, "D": 2}
    rows = [
        (13, "TotCust", 16), (14, "MRR", 16), (15, "ARR", 16), (16, "Cash", 16),
        (17, "GrossRev", (5, 16)), (18, "NetRev", (5, 16)), (19, "Profit", (5, 16)),
        (20, "TotCust", 28), (21, "MRR", 28), (22, "ARR", 28), (23, "Cash", 28),
        (24, "GrossRev", (17, 28)),
        (25, "TotCust", 64), (26, "MRR", 64), (27, "ARR", 64), (28, "Cash", 64),
    ]
    for col, block in blocks.items():
        for row, metric, target in rows:
            if isinstance(target, tuple):
                ws[f"{col}{row}"] = se_sum(block, metric, target[0], target[1])
            else:
                ws[f"{col}{row}"] = f"={se_ref(block, metric, target)}"


def update_sens_churn(wb):
    ws = wb["Sens - Churn"]
    ws["A1"] = "Churn Sensitivity - v3 Elite full model"
    ws["B5"] = model_ref("MRR", 16)
    ws["B6"] = model_ref("MRR", 28)
    ws["B7"] = model_ref("ARR", 28)
    ws["B8"] = model_ref("ARR", 64)
    ws["B9"] = model_ref("Cash", 64)
    for i in range(9):
        r = 13 + i
        block = 3 + i
        ws.cell(r, 2, f"={se_ref(block, 'MRR', 16)}")
        ws.cell(r, 3, f"={se_ref(block, 'TotCust', 16)}")
        ws.cell(r, 4, f"={se_ref(block, 'MRR', 28)}")
        ws.cell(r, 5, f"={se_ref(block, 'ARR', 28)}")


def update_sens_annual_mix(wb):
    ws = wb["Sens - Annual Mix"]
    ws["A1"] = "Annual Plan Mix Sensitivity - v3 Elite engine"
    for i in range(7):
        r = 9 + i
        block = 12 + i
        cust = se_ref(block, "TotCust", 16)
        mrr = se_ref(block, "MRR", 16)
        ws.cell(r, 2, f"={mrr}")
        ws.cell(r, 3, f"=IF({cust}>0,{mrr}/{cust},0)")
        ws.cell(r, 4, f"={se_ref(block, 'MRR', 28)}")
        ws.cell(r, 5, f"={se_ref(block, 'ARR', 28)}")


def update_sens_cac(wb):
    ws = wb["Sens - CAC & Marketing"]
    ws["A1"] = "CAC Sensitivity - v3 Elite full model engine"
    for i in range(7):
        r = 9 + i
        block = 19 + i
        ws.cell(r, 2, se_sum(block, "PaidNew", 5, 16))
        ws.cell(r, 3, f"={se_ref(block, 'TotCust', 16)}")
        ws.cell(r, 4, f"={se_ref(block, 'MRR', 16)}")
        ws.cell(r, 5, f"={se_ref(block, 'MRR', 28)}")
        ws.cell(r, 6, f"={se_ref(block, 'Cash', 28)}")
        ws.cell(r, 7, f"={se_ref(block, 'Cash', 64)}")


def update_sens_winter(wb, base_refs, winter_refs, season_start):
    ws = wb["Sens - Winter Pivot"]
    ws["A1"] = "Winter Pivot - Property Care Scenario (v3 Elite)"
    for i, mo in enumerate(range(1, 13)):
        r = 13 + i
        ir = season_start + mo - 1
        ws.cell(r, 2, f"=Inputs!B{ir}")
        ws.cell(r, 3, f"=Inputs!E{ir}")
        ws.cell(r, 5, f"=Inputs!D{ir}")
        ws.cell(r, 6, f"=Inputs!F{ir}")
    metrics = [
        ("End Year 1 - Customers", 16), ("End Year 1 - MRR (£)", 16), ("End Year 1 - ARR (£)", 16),
        ("End Year 1 - Cash (£)", 16), ("Year 1 - Gross Revenue (£)", (5, 16)), ("Year 1 - Profit (£)", (5, 16)),
        ("End Year 2 - Customers", 28), ("End Year 2 - MRR (£)", 28), ("End Year 2 - Cash (£)", 28),
        ("End Year 5 - Customers", 64), ("End Year 5 - MRR (£)", 64), ("End Year 5 - Cash (£)", 64),
    ]
    for i, (label, target) in enumerate(metrics):
        r = 31 + i
        ws.cell(r, 1, label)
        if isinstance(target, tuple):
            ws.cell(r, 2, se_sum(1, "GrossRev" if "Gross" in label else "Profit", target[0], target[1]))
            ws.cell(r, 3, se_sum(26, "GrossRev" if "Gross" in label else "Profit", target[0], target[1]))
        else:
            metric = "TotCust" if "Customers" in label else "MRR" if "MRR" in label else "ARR" if "ARR" in label else "Cash"
            ws.cell(r, 2, f"={se_ref(1, metric, target)}")
            ws.cell(r, 3, f"={se_ref(26, metric, target)}")
        ws.cell(r, 4, f"=C{r}-B{r}")


def rebuild_model_and_scenarios(wb):
    idx = wb.sheetnames.index("Model")
    del wb["Model"]
    ws_m = wb.create_sheet("Model", idx)
    ws_m["A1"] = "60-Month Model v4 - 10/20/30 visits per year + VAT principal/agent toggle"
    ws_m["A1"].font = TITLE_FONT
    M = write_block(ws_m, 1, None, "MonthlyChurn", "AnnualChurn", "AnnualMix", "CAC")

    for m, lb in [(12, "Month12"), (24, "Month24"), (60, "Month60")]:
        er = FIRST + m - 1
        for nm, cn in [("MRR", "MRR"), ("ARR", "ARR"), ("Cust", "TotCust"), ("Cash", "Cash")]:
            set_defined_name(wb, f"{nm}_{lb}", f"Model!${M[cn]}${er}")

    if "Scenario Engine" in wb.sheetnames:
        del wb["Scenario Engine"]
    ws_se = wb.create_sheet("Scenario Engine")
    ws_se.sheet_state = "hidden"

    blocks = {}
    blocks["cons"] = write_block(ws_se, block_sc(0), "Conservative", "'Sens - Scenarios'!$B$5", "'Sens - Scenarios'!$B$6", "'Sens - Scenarios'!$B$7", "'Sens - Scenarios'!$B$8", hdr_row=3)
    blocks["base"] = write_block(ws_se, block_sc(1), "Base", "MonthlyChurn", "AnnualChurn", "AnnualMix", "CAC", hdr_row=3)
    blocks["agg"] = write_block(ws_se, block_sc(2), "Aggressive", "'Sens - Scenarios'!$D$5", "'Sens - Scenarios'!$D$6", "'Sens - Scenarios'!$D$7", "'Sens - Scenarios'!$D$8", hdr_row=3)
    for i in range(9):
        blocks[f"ch{i}"] = write_block(ws_se, block_sc(3 + i), f"Ch{i}", f"'Sens - Churn'!$A${13 + i}", "AnnualChurn", "AnnualMix", "CAC", hdr_row=3)
    for i in range(7):
        blocks[f"mx{i}"] = write_block(ws_se, block_sc(12 + i), f"Mx{i}", "MonthlyChurn", "AnnualChurn", f"'Sens - Annual Mix'!$A${9 + i}", "CAC", hdr_row=3)
    for i in range(7):
        blocks[f"cac{i}"] = write_block(ws_se, block_sc(19 + i), f"Cac{i}", "MonthlyChurn", "AnnualChurn", "AnnualMix", f"'Sens - CAC & Marketing'!$A${9 + i}", hdr_row=3)

    winter_sc = block_sc(26)
    for r in range(1, 65):
        for c in range(winter_sc, winter_sc + BLOCK_W):
            ws_se.cell(r, c).value = None
    WP = write_winter_block(
        ws_se, winter_sc, "Winter Pivot",
        essmix="'Sens - Winter Pivot'!$B$6",
        prem_uplift="'Sens - Winter Pivot'!$B$7",
    )
    return blocks["base"], WP


def main():
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    season_start = update_inputs(wb)
    update_assumptions(wb)
    base_refs, winter_refs = rebuild_model_and_scenarios(wb)
    update_headlines(wb)
    update_five_year_summary(wb)
    update_24_month_forecast(wb)
    update_unit_economics(wb)
    update_sens_scenarios(wb)
    update_sens_churn(wb)
    update_sens_annual_mix(wb)
    update_sens_cac(wb)
    update_sens_winter(wb, base_refs, winter_refs, season_start)
    write_vat_dual_calc(wb)
    update_vat_comparison(wb)
    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Model TotCust col: {cl(1, 'TotCust')}, MRR: {cl(1, 'MRR')}, Cash: {cl(1, 'Cash')}")


if __name__ == "__main__":
    main()
