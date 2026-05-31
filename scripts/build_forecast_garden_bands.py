#!/usr/bin/env python3
"""Build simplified SaaS forecast with 4 garden bands and UK-derived customer mix."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "sorted_saas_forecast_garden_bands.xlsx"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
NOTE_FONT = Font(italic=True, color="666666", size=10)
MONEY = "£#,##0.00"
MONEY0 = "£#,##0"
PCT = "0.0%"

ESSENTIAL_VISITS_PER_YEAR = 10
TRAVEL_MINUTES = 10

# (name, max m², customer £/mo Essential, provider £/visit, on-site min, slot min)
# Format: garden size - provider time - price/mo - provider pay/visit
BANDS = [
    ("Small", "≤50 m²", 59.99, 20.00, 60, 70),
    ("Medium", "≤100 m²", 79.99, 30.00, 90, 100),
    ("Large", "≤150 m²", 99.99, 40.00, 120, 130),
]

UK_HOUSING_MIX = (0.25, 0.45, 0.30)

# Subscriber mix (edit on sheet) - default centres on Medium band
FORECAST_MIX = (0.15, 0.50, 0.35)
TARGET_BLENDED_PRICE = sum(BANDS[i][2] * FORECAST_MIX[i] for i in range(3))

FIRST_MODEL_ROW = 5
MODEL_MONTHS = 60


def style_input(ws, row, col=2):
    c = ws.cell(row, col)
    c.fill = INPUT_FILL
    c.font = Font(size=11)


def style_money(ws, row, col, fmt=MONEY):
    c = ws.cell(row, col)
    c.number_format = fmt


def style_pct(ws, row, col):
    ws.cell(row, col).number_format = PCT


def build_garden_mix(wb):
    ws = wb.create_sheet("Garden mix", 1)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 40

    ws["A1"] = "UK garden band mix (maintained lawn & beds)"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Band"
    ws["B3"] = "Max m²"
    ws["C3"] = "Mix %"
    ws["D3"] = "Typical UK fit"
    for c in range(1, 5):
        cell = ws.cell(3, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # UK mapping table
    ws["A5"] = "HOW THE MIX IS BUILT (England homes with a garden)"
    ws["A5"].font = SECTION_FONT
    rows = [
        ("Dwelling type", "Share", "Maintained m²", "Band split"),
        ("Mid-terrace rear", 0.22, 30, "100% Courtyard"),
        ("End-terrace rear", 0.12, 48, "60% Courtyard / 40% Standard"),
        ("Semi-detached rear", 0.32, 72, "60% Standard / 40% Family"),
        ("Detached rear", 0.24, 95, "55% Family / 45% Large"),
        ("Bungalow", 0.10, 65, "70% Standard / 30% Family"),
    ]
    r = 6
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        r += 1
    ws["A12"] = (
        "Maintained area ≈ lawn + beds we cut (not whole plot). "
        "Sources: JRF new-build rear sizes; adjusted for existing stock & paved areas. "
        ">140 m² (~2%) excluded - quoted separately; mix renormalised to 100%."
    )
    ws["A12"].font = NOTE_FONT
    ws.merge_cells("A12:D12")

    ws["A14"] = (
        f"FORECAST MIX - Essential, {ESSENTIAL_VISITS_PER_YEAR} visits/yr "
        f"(edit yellow mix %; provider £/mo = £/visit × {ESSENTIAL_VISITS_PER_YEAR}/12)"
    )
    ws["A14"].font = SECTION_FONT
    ws.merge_cells("A14:J14")

    headers = [
        (1, "Band"),
        (2, "Max m²"),
        (3, "Customer £/mo"),
        (4, "Provider £/visit"),
        (5, "Provider £/mo"),
        (6, "On-site min"),
        (7, f"Slot min (+{TRAVEL_MINUTES} travel)"),
        (8, "Forecast mix"),
        (9, "UK housing mix"),
        (10, "Notes"),
    ]
    for col, label in headers:
        cell = ws.cell(15, col, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, (name, area, price, prov_visit, onsite, slot) in enumerate(BANDS):
        r = 16 + i
        visit_hrs = onsite / 60
        ws.cell(r, 1, name)
        ws.cell(r, 2, area)
        ws.cell(r, 3, price)
        style_money(ws, r, 3)
        ws.cell(r, 4, prov_visit)
        style_money(ws, r, 4)
        ws.cell(r, 5, f"=D{r}*{ESSENTIAL_VISITS_PER_YEAR}/12")
        style_money(ws, r, 5)
        ws.cell(r, 6, onsite)
        ws.cell(r, 7, slot)
        ws.cell(r, 8, FORECAST_MIX[i])
        style_pct(ws, r, 8)
        style_input(ws, r, 8)
        ws.cell(r, 9, UK_HOUSING_MIX[i])
        style_pct(ws, r, 9)
        ws.cell(
            r,
            10,
            f"{visit_hrs:g} hr on site · £{prov_visit:.0f}/visit · £{price:.2f}/mo Essential",
        )
        ws.cell(r, 10).font = NOTE_FONT

    ws["A20"] = "Forecast mix sums to 100%"
    ws.cell(20, 8, "=SUM(H16:H18)")
    style_pct(ws, 20, 8)

    ws["A21"] = "UK housing mix (reference only)"
    ws.cell(21, 8, "=SUM(I16:I18)")
    style_pct(ws, 21, 8)

    ws["A22"] = "Blended customer price (→ Headlines)"
    ws.cell(22, 3, "=SUMPRODUCT(C16:C18,H16:H18)")
    style_money(ws, 22, 3)
    ws.cell(22, 10, f"Target ~£{TARGET_BLENDED_PRICE:.2f} Essential monthly")
    ws.cell(22, 10).font = NOTE_FONT

    ws["A23"] = "Blended price - UK housing mix only"
    ws.cell(23, 3, "=SUMPRODUCT(C16:C18,I16:I18)")
    style_money(ws, 23, 3)
    ws.cell(23, 10, "Typical homes undershoot paid subscribers")
    ws.cell(23, 10).font = NOTE_FONT

    ws["A24"] = "Blended provider £/visit"
    ws.cell(24, 4, "=SUMPRODUCT(D16:D18,H16:H18)")
    style_money(ws, 24, 4)

    ws["A25"] = "Blended provider £/mo (→ Headlines)"
    ws.cell(25, 5, "=SUMPRODUCT(E16:E18,H16:H18)")
    style_money(ws, 25, 5)

    ws["A26"] = "Blended on-site time (min)"
    ws.cell(26, 6, "=ROUND(SUMPRODUCT(F16:F18,H16:H18),0)")
    ws.cell(26, 10, "Weighted avg Essential visit duration")
    ws.cell(26, 10).font = NOTE_FONT

    ws["A27"] = "Blended slot time (min)"
    ws.cell(27, 7, "=ROUND(SUMPRODUCT(G16:G18,H16:H18),0)")
    ws.cell(27, 10, f"Incl. {TRAVEL_MINUTES} min travel between jobs")
    ws.cell(27, 10).font = NOTE_FONT

    ws["A29"] = "Implied contribution (pre-fees, pre-ops)"
    ws.cell(29, 3, "=C22-E25-5")
    style_money(ws, 29, 3)
    ws.cell(29, 10, "Assumes £5 ops/customer - change on Headlines")
    ws.cell(29, 10).font = NOTE_FONT

    ws["A31"] = (
        "Why two mixes? Housing stock skews small (blended ~£52 customer price). "
        "People who buy maintenance skew larger - use forecast mix for cash planning."
    )
    ws["A31"].font = NOTE_FONT
    ws.merge_cells("A31:J31")


def build_headlines(wb):
    ws = wb.active
    ws.title = "Headlines"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["D"].width = 42

    ws["A1"] = "GardensSorted - Forecast (4 garden bands)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Edit yellow cells. Blended price & provider come from Garden mix tab."
    ws["A2"].font = NOTE_FONT

    ws["A4"] = "INPUTS"
    ws["A4"].font = SECTION_FONT

    inputs = [
        (5, "Initial investment (£)", 17000, "Starting cash"),
        (7, "CAC (£)", 70, "Paid acquisition per customer"),
        (8, "Monthly churn (%/mo)", 0.055, "After 3-month minimum"),
        (9, "Payment processing (%)", 0.03, "Stripe/card"),
        (10, "Ops per customer (£/mo)", 5, "Support, hosting, AI"),
        (11, "Flat platform cost (£/mo)", 250, "Tools & hosting"),
        (12, "Marketing spend (£/mo)", 3000, "Flat budget"),
        (14, "VAT rate (%)", 0.2, "Once registered"),
        (15, "VAT threshold (£ rolling 12m)", 90000, None),
        (16, "Annual inflation (%)", 0.025, "Price uplift per model year"),
        (17, "Launch calendar month (1=Jan)", 3, "March pre-season"),
    ]
    for row, label, val, note in inputs:
        ws.cell(row, 1, label)
        ws.cell(row, 2, val)
        style_input(ws, row)
        if note:
            ws.cell(row, 4, note)
            ws.cell(row, 4).font = NOTE_FONT
        if row in (8, 9, 14, 16):
            style_pct(ws, row, 2)

    ws["A19"] = "FROM GARDEN MIX (calculated)"
    ws["A19"].font = SECTION_FONT
    ws["A20"] = "Blended plan price (£/mo)"
    ws.cell(20, 2, "='Garden mix'!C22")
    style_money(ws, 20, 2)
    ws["A21"] = "Blended provider pay (£/mo)"
    ws.cell(21, 2, "='Garden mix'!E25")
    style_money(ws, 21, 2)

    ws["A23"] = "KEY METRICS"
    ws["A23"].font = SECTION_FONT
    ws["A24"] = "Metric"
    ws["B24"] = "Year 1"
    ws["C24"] = "Year 2"
    ws["D24"] = "Year 5"
    metrics = [
        ("End customers", "=Model!I16", "=Model!I28", "=Model!I64"),
        ("MRR", "=Model!J16", "=Model!J28", "=Model!J64"),
        ("ARR", "=Model!K16", "=Model!K28", "=Model!K64"),
        ("Gross revenue", "=SUM(Model!L5:L16)", "=SUM(Model!L17:L28)", "=SUM(Model!L53:L64)"),
        ("Net revenue", "=SUM(Model!O5:O16)", "=SUM(Model!O17:O28)", "=SUM(Model!O53:O64)"),
        ("Annual profit", "=SUM(Model!W5:W16)", "=SUM(Model!W17:W28)", "=SUM(Model!W53:W64)"),
        ("Cash balance", "=Model!X16", "=Model!X28", "=Model!X64"),
    ]
    for idx, (name, y1, y2, y5) in enumerate(metrics):
        r = 25 + idx
        ws.cell(r, 1, name)
        ws.cell(r, 2, y1)
        ws.cell(r, 3, y2)
        ws.cell(r, 4, y5)
        if name != "End customers":
            for col in range(2, 5):
                style_money(ws, r, col, MONEY0)

    ws["A33"] = "MILESTONES"
    ws["A33"].font = SECTION_FONT
    ws["A34"] = "VAT registration"
    ws.cell(34, 2, '=IFERROR("Month "&MINIFS(Model!A5:A64,Model!N5:N64,1),"Not in 60 months")')
    ws["A35"] = "First profitable month"
    ws.cell(35, 2, '=IFERROR("Month "&MINIFS(Model!A5:A64,Model!W5:W64,">0"),"Not in 60 months")')
    ws["A36"] = "Cash payback"
    ws.cell(36, 2, '=IFERROR("Month "&MINIFS(Model!A5:A64,Model!X5:X64,">="&B5),"Not reached")')

    ws["A38"] = "SEASONALITY"
    ws["A38"].font = SECTION_FONT
    ws["A39"] = "Month"
    ws["B39"] = "Acq mult"
    ws["C39"] = "Churn mult"
    season = [
        (1, 0.56, 1.37),
        (2, 0.66, 1.28),
        (3, 0.87, 1.04),
        (4, 1.07, 0.85),
        (5, 1.17, 0.76),
        (6, 1.22, 0.71),
        (7, 1.22, 0.71),
        (8, 1.17, 0.76),
        (9, 1.02, 0.90),
        (10, 0.82, 1.04),
        (11, 0.61, 1.23),
        (12, 0.56, 1.37),
    ]
    for i, (m, acq, ch) in enumerate(season, 40):
        ws.cell(i, 1, m)
        ws.cell(i, 2, acq)
        ws.cell(i, 3, ch)


def col_letter(n):
    return get_column_letter(n)


def build_model(wb):
    ws = wb.create_sheet("Model")
    headers = [
        "Mo",
        "Cal",
        "CY",
        "Acq",
        "Infl",
        "ECh",
        "Mkt",
        "New",
        "Cust",
        "MRR",
        "ARR",
        "Gross",
        "R12",
        "VAT?",
        "Net",
        "VAcc",
        "VPay",
        "Fees",
        "Prov",
        "Ops",
        "Fixed",
        "Mkt$",
        "Profit",
        "Cash",
        "Chrn",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    launch = "Headlines!$B$17"
    cac = "Headlines!$B$7"
    churn = "Headlines!$B$8"
    price = "Headlines!$B$20"
    prov = "Headlines!$B$21"
    fee = "Headlines!$B$9"
    ops = "Headlines!$B$10"
    fixed = "Headlines!$B$11"
    mkt = "Headlines!$B$12"
    infl = "Headlines!$B$16"
    vat_r = "Headlines!$B$14"
    vat_th = "Headlines!$B$15"
    invest = "Headlines!$B$5"
    season = "Headlines!$A$40:$C$51"

    for i in range(MODEL_MONTHS):
        r = FIRST_MODEL_ROW + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, f"=MOD(A{r}-1+{launch}-1,12)+1")
        ws.cell(r, 3, f"=INT((A{r}-1+{launch}-1)/12)+1")
        ws.cell(r, 4, f"=VLOOKUP(B{r},{season},2,FALSE)")
        ws.cell(r, 5, f"=(1+{infl})^INT((A{r}-1)/12)")
        ws.cell(r, 6, f"={churn}*VLOOKUP(B{r},{season},3,FALSE)")
        ws.cell(r, 7, f"={mkt}")

        ws.cell(r, 8, f"=IF(G{r}>0,MAX(1,ROUND(G{r}/{cac}*D{r},0)),0)")

        if r == FIRST_MODEL_ROW:
            cust = f"=H{r}"
        elif r <= FIRST_MODEL_ROW + 3:
            cust = f"=I{r-1}+H{r}"
        else:
            cust = (
                f"=H{r}+H{r-1}+H{r-2}+MAX(0,I{r-1}-H{r-1}-H{r-2})*(1-F{r})"
            )
        ws.cell(r, 9, cust)

        ws.cell(r, 10, f"=I{r}*{price}*E{r}")
        ws.cell(r, 11, f"=J{r}*12")
        ws.cell(r, 12, f"=J{r}")
        ws.cell(r, 13, f"=SUM(L$5:L{r})")
        ws.cell(r, 14, f"=IF(M{r}>={vat_th},1,0)")
        ws.cell(r, 15, f"=IF(N{r},L{r}/(1+{vat_r}),L{r})")
        ws.cell(r, 16, f"=L{r}-O{r}")

        # Quarterly VAT payment (simplified)
        b = f"B{r}"
        cy = f"C{r}"
        vp = (
            f"=IF(N{r}=0,0,IF({b}=1,SUMIFS(P$5:P${r},C$5:C${r},{cy}-1,{b}$5:{b}${r},\">=10\",{b}$5:{b}${r},\"<=12\"),"
            f"IF({b}=4,SUMIFS(P$5:P${r},C$5:C${r},{cy},{b}$5:{b}${r},\">=1\",{b}$5:{b}${r},\"<=3\"),"
            f"IF({b}=7,SUMIFS(P$5:P${r},C$5:C${r},{cy},{b}$5:{b}${r},\">=4\",{b}$5:{b}${r},\"<=6\"),"
            f"IF({b}=10,SUMIFS(P$5:P${r},C$5:C${r},{cy},{b}$5:{b}${r},\">=7\",{b}$5:{b}${r},\"<=9\"),0)))))"
        )
        ws.cell(r, 17, vp)

        ws.cell(r, 18, f"=L{r}*{fee}")
        ws.cell(r, 19, f"=I{r}*{prov}")
        ws.cell(r, 20, f"=I{r}*{ops}")
        ws.cell(r, 21, f"={fixed}")
        ws.cell(r, 22, f"=G{r}")
        ws.cell(r, 23, f"=O{r}-R{r}-S{r}-T{r}-U{r}-V{r}")
        if r == FIRST_MODEL_ROW:
            ws.cell(r, 24, f"={invest}+L{r}-(R{r}+S{r}+T{r}+U{r}+V{r}+Q{r})")
        else:
            ws.cell(r, 24, f"=X{r-1}+L{r}-(R{r}+S{r}+T{r}+U{r}+V{r}+Q{r})")
        ws.cell(r, 25, f"=F{r}")

        for c in (10, 11, 12, 15, 16, 18, 19, 20, 23, 24):
            style_money(ws, r, c, MONEY0)
        style_money(ws, r, 13, MONEY0)
        style_money(ws, r, 17, MONEY0)


def build_summary(wb):
    ws = wb.create_sheet("5 Year Summary")
    ws["A1"] = "5-Year Summary"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Year"
    ws["B3"] = "Customers"
    ws["C3"] = "MRR"
    ws["D3"] = "ARR"
    ws["E3"] = "New"
    ws["F3"] = "Gross"
    ws["G3"] = "Net"
    for c in range(1, 8):
        ws.cell(3, c).fill = HEADER_FILL
        ws.cell(3, c).font = HEADER_FONT
    years = [(1, 16), (2, 28), (3, 40), (4, 52), (5, 64)]
    for i, (yr, end_row) in enumerate(years, 4):
        start = end_row - 11
        ws.cell(i, 1, f"Year {yr}")
        ws.cell(i, 2, f"=Model!I{end_row}")
        ws.cell(i, 3, f"=Model!J{end_row}")
        ws.cell(i, 4, f"=Model!K{end_row}")
        ws.cell(i, 5, f"=SUM(Model!H{start}:H{end_row})")
        ws.cell(i, 6, f"=SUM(Model!L{start}:L{end_row})")
        ws.cell(i, 7, f"=SUM(Model!O{start}:O{end_row})")
        for c in range(3, 8):
            style_money(ws, i, c, MONEY0)


def main():
    wb = openpyxl.Workbook()
    build_headlines(wb)
    build_garden_mix(wb)
    build_model(wb)
    build_summary(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    n = len(BANDS)
    b_price = sum(BANDS[i][2] * FORECAST_MIX[i] for i in range(n))
    b_prov_visit = sum(BANDS[i][3] * FORECAST_MIX[i] for i in range(n))
    b_prov_mo = b_prov_visit * (ESSENTIAL_VISITS_PER_YEAR / 12)
    b_onsite = sum(BANDS[i][4] * FORECAST_MIX[i] for i in range(n))
    b_slot = sum(BANDS[i][5] * FORECAST_MIX[i] for i in range(n))
    uk_price = sum(BANDS[i][2] * UK_HOUSING_MIX[i] for i in range(n))
    print(f"  Forecast mix: {FORECAST_MIX}")
    print(f"  Blended price: £{b_price:.2f}/mo  Provider: £{b_prov_visit:.2f}/visit  £{b_prov_mo:.2f}/mo")
    print(f"  Blended time: {b_onsite:.0f} min on-site, {b_slot:.0f} min slot")
    print(f"  UK housing mix price (ref): £{uk_price:.2f}/mo")


if __name__ == "__main__":
    main()
