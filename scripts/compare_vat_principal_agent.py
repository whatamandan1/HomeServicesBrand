#!/usr/bin/env python3
"""Compare principal vs agent VAT impact — mirrors v4 model base case (10/20/30 visits per year)."""
import openpyxl

XLSX = "/Users/dan/Documents/HomeServicesBrand/sorted_saas_recurring_revenue_forecast_v3_elite.xlsx"
FIRST = 5

VISITS_ESS_YR = 10
VISITS_PREM_YR = 20
VISITS_ELITE_YR = 30
PROV_SMALL = 15
PROV_MEDIUM = 18
PROV_LARGE = 21


def load_inputs(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    wf = openpyxl.load_workbook(path, data_only=False)

    def named(name):
        ref = wf.defined_names[name].attr_text
        sheet, cell = ref.split("!")
        return float(wb[sheet.strip("'")][cell.replace("$", "")].value)

    season_start = 74
    ws = wb["Inputs"]
    acq = {}
    churn = {}
    for mo in range(1, 13):
        r = season_start + mo - 1
        acq[mo] = float(ws.cell(r, 2).value)
        churn[mo] = float(ws.cell(r, 4).value)

    small = named("GardenSmallMix")
    med = named("GardenMediumMix")
    large = named("GardenLargeMix")

    def blend_price(base, mo_up, ann_up, annual=False):
        up = ann_up if annual else mo_up
        return base * small + (base + up) * med + (base + up * (2 if not annual else 10)) * large

    # monthly uplifts +10/+20; annual +100/+200
    prices = {
        "ess_m": blend_price(named("PriceEssMonthly"), 10, 100),
        "prem_m": blend_price(named("PricePremMonthly"), 10, 100),
        "elite_m": blend_price(named("PriceEliteMonthly"), 10, 100),
        "ess_a": named("PriceEssAnnual") * small + (named("PriceEssAnnual") + 100) * med + (named("PriceEssAnnual") + 200) * large,
        "prem_a": named("PricePremAnnual") * small + (named("PricePremAnnual") + 100) * med + (named("PricePremAnnual") + 200) * large,
        "elite_a": named("PriceEliteAnnual") * small + (named("PriceEliteAnnual") + 100) * med + (named("PriceEliteAnnual") + 200) * large,
    }
    pay = {
        "ess": (VISITS_ESS_YR / 12) * (small * PROV_SMALL + med * PROV_MEDIUM + large * PROV_LARGE),
        "prem": (VISITS_PREM_YR / 12) * (small * PROV_SMALL + med * PROV_MEDIUM + large * PROV_LARGE),
        "elite": (VISITS_ELITE_YR / 12) * (small * PROV_SMALL + med * PROV_MEDIUM + large * PROV_LARGE),
    }

    return {
        "annmix": named("AnnualMix"),
        "essmix": named("EssentialMix"),
        "premmix": named("PremiumMix"),
        "mchurn": named("MonthlyChurn"),
        "achurn": named("AnnualChurn"),
        "cac": named("CAC"),
        "pfee": named("PaymentFee"),
        "ops": named("OpsAI"),
        "fixed": named("MonthlyFixed"),
        "pdelay": int(named("ProviderDelay")),
        "init": named("InitialInvestment"),
        "buffer": named("CashBuffer"),
        "refmix": named("ReferralMix"),
        "refreward": named("ReferralReward"),
        "infl": named("InflationRate"),
        "vat_rate": named("VATRate"),
        "vat_thresh": named("VATThreshold"),
        "mkt13": named("MktFixed1_3"),
        "mkt46": named("MktFixed4_6"),
        "pct712": named("MktPct7_12"),
        "pct1318": named("MktPct13_18"),
        "pct19": named("MktPct19plus"),
        "min712": named("MktMin7_12"),
        "min1318": named("MktMin13_18"),
        "min19": named("MktMin19plus"),
        "launch": int(named("LaunchMonth")),
        "acq": acq,
        "churn_m": churn,
        **prices,
        **{f"pay_{k}": v for k, v in pay.items()},
    }


def monthly_active(new_series, eff_churn):
    """3-month min term cohort roll-forward."""
    active = 0.0
    history = []
    for m, new in enumerate(new_series, 1):
        history.append(new)
        if m == 1:
            active = new
        elif m == 2:
            active = history[0] + history[1]
        elif m == 3:
            active = sum(history)
        else:
            active = sum(history[-3:]) + max(0, active - sum(history[-4:-1])) * (1 - eff_churn)
        yield active


def simulate(cfg):
    rows = []
    vat_reg_p = vat_reg_a = False
    cash = cfg["init"]

    # Track monthly new subs by tier for monthly cohorts
    new_es_m = []
    new_pr_m = []
    new_el_m = []
    new_es_a = []
    new_pr_a = []
    new_el_a = []

    act_es_m = act_pr_m = act_el_m = 0.0
    act_es_a = act_pr_a = act_el_a = 0.0

    gross_hist = []
    margin_hist = []

    for m in range(1, 61):
        calmo = (m - 1 + cfg["launch"] - 1) % 12 + 1
        infl = (1 + cfg["infl"]) ** ((m - 1) // 12)
        eff_ch = cfg["mchurn"] * cfg["churn_m"][calmo]
        acq = cfg["acq"][calmo]

        gross_prev = rows[-1]["gross"] if rows else 0.0
        cb = cash
        if m <= 3:
            mkt = cfg["mkt13"]
        elif m <= 6:
            mkt = min(cfg["mkt46"], max(0, cb - cfg["buffer"]))
        elif m <= 12:
            mkt = min(max(cfg["min712"], gross_prev * cfg["pct712"]), max(0, cb - cfg["buffer"]))
        elif m <= 18:
            mkt = min(max(cfg["min1318"], gross_prev * cfg["pct1318"]), max(0, cb - cfg["buffer"]))
        else:
            mkt = min(max(cfg["min19"], gross_prev * cfg["pct19"]), max(0, cb - cfg["buffer"]))

        paid = max(1, round(mkt / cfg["cac"] * acq)) if mkt > 0 else 0
        ref = round(paid * cfg["refmix"] / (1 - cfg["refmix"])) if paid > 0 else 0
        tot = paid + ref
        nmo = round(tot * (1 - cfg["annmix"]))
        nan = tot - nmo
        nesm = round(nmo * cfg["essmix"])
        nprm = round(nmo * cfg["premmix"])
        nelm = nmo - nesm - nprm
        nesa = round(nan * cfg["essmix"])
        npra = round(nan * cfg["premmix"])
        nela = nan - nesa - npra

        new_es_m.append(nesm)
        new_pr_m.append(nprm)
        new_el_m.append(nelm)

        if m == 1:
            act_es_m = nesm
            act_pr_m = nprm
            act_el_m = nelm
        elif m == 2:
            act_es_m = new_es_m[-1] + new_es_m[-2]
            act_pr_m = new_pr_m[-1] + new_pr_m[-2]
            act_el_m = new_el_m[-1] + new_el_m[-2]
        elif m == 3:
            act_es_m = sum(new_es_m[-3:])
            act_pr_m = sum(new_pr_m[-3:])
            act_el_m = sum(new_el_m[-3:])
        else:
            p = rows[-1]
            act_es_m = sum(new_es_m[-3:]) + max(0, p["act_es_m"] - sum(new_es_m[-4:-1])) * (1 - eff_ch)
            act_pr_m = sum(new_pr_m[-3:]) + max(0, p["act_pr_m"] - sum(new_pr_m[-4:-1])) * (1 - eff_ch)
            act_el_m = sum(new_el_m[-3:]) + max(0, p["act_el_m"] - sum(new_el_m[-4:-1])) * (1 - eff_ch)

        new_es_a.append(nesa)
        new_pr_a.append(npra)
        new_el_a.append(nela)
        if m <= 12:
            act_es_a = sum(new_es_a)
            act_pr_a = sum(new_pr_a)
            act_el_a = sum(new_el_a)
        else:
            # simplified annual renewal churn
            act_es_a = rows[-1]["act_es_a"] + nesa - rows[-1]["act_es_a"] * cfg["achurn"] * 0.08
            act_pr_a = rows[-1]["act_pr_a"] + npra - rows[-1]["act_pr_a"] * cfg["achurn"] * 0.08
            act_el_a = rows[-1]["act_el_a"] + nela - rows[-1]["act_el_a"] * cfg["achurn"] * 0.08

        mobill = (
            act_es_m * cfg["ess_m"] + act_pr_m * cfg["prem_m"] + act_el_m * cfg["elite_m"]
        ) * infl
        newann = (nesa * cfg["ess_a"] + npra * cfg["prem_a"] + nela * cfg["elite_a"]) * infl
        ren = 0.0
        gross = mobill + newann + ren

        prov_now = (
            act_es_m * cfg["pay_ess"] + act_pr_m * cfg["pay_prem"] + act_el_m * cfg["pay_elite"]
            + act_es_a * cfg["pay_ess"] + act_pr_a * cfg["pay_prem"] + act_el_a * cfg["pay_elite"]
        )
        if cfg["pdelay"] >= 1:
            prov = rows[-1]["prov_now"] if rows else 0.0
        else:
            prov = prov_now

        margin = gross - prov
        gross_hist.append(gross)
        margin_hist.append(margin)
        roll12_g = sum(gross_hist[-12:])
        roll12_m = sum(margin_hist[-12:])

        if not vat_reg_p:
            vat_reg_p = roll12_g >= cfg["vat_thresh"]
        if not vat_reg_a:
            vat_reg_a = roll12_m >= cfg["vat_thresh"]

        vr = cfg["vat_rate"]
        net_p = gross / (1 + vr) if vat_reg_p else gross
        if vat_reg_a:
            vat_on_margin = margin - margin / (1 + vr)
            net_a = gross - vat_on_margin
        else:
            net_a = gross

        cust = act_es_m + act_pr_m + act_el_m + act_es_a + act_pr_a + act_el_a
        fees = gross * cfg["pfee"]
        ops = cust * cfg["ops"]
        refc = ref * cfg["refreward"]
        profit_p = net_p - prov - fees - ops - cfg["fixed"] - refc - mkt
        profit_a = net_a - prov - fees - ops - cfg["fixed"] - refc - mkt

        rows.append(
            {
                "m": m,
                "year": (m - 1) // 12 + 1,
                "gross": gross,
                "prov": prov,
                "prov_now": prov_now,
                "margin": margin,
                "net_p": net_p,
                "net_a": net_a,
                "gp_p": net_p - prov,
                "gp_a": net_a - prov,
                "profit_p": profit_p,
                "profit_a": profit_a,
                "vat_reg_p": vat_reg_p,
                "vat_reg_a": vat_reg_a,
                "act_es_m": act_es_m,
                "act_pr_m": act_pr_m,
                "act_el_m": act_el_m,
                "act_es_a": act_es_a,
                "act_pr_a": act_pr_a,
                "act_el_a": act_el_a,
            }
        )
        cash = cash + gross - prov - fees - ops - cfg["fixed"] - refc - mkt

    return rows


def year_totals(rows, year):
    chunk = [r for r in rows if r["year"] == year]
    if not chunk:
        return None
    return {
        "gross": sum(r["gross"] for r in chunk),
        "prov": sum(r["prov"] for r in chunk),
        "net_p": sum(r["net_p"] for r in chunk),
        "net_a": sum(r["net_a"] for r in chunk),
        "gp_p": sum(r["gp_p"] for r in chunk),
        "gp_a": sum(r["gp_a"] for r in chunk),
        "profit_p": sum(r["profit_p"] for r in chunk),
        "profit_a": sum(r["profit_a"] for r in chunk),
        "vat_p": sum(r["gross"] - r["net_p"] for r in chunk),
        "vat_a": sum(r["gross"] - r["net_a"] for r in chunk),
        "vat_reg_p": chunk[-1]["vat_reg_p"],
        "vat_reg_a": chunk[-1]["vat_reg_a"],
        "cust": chunk[-1]["act_es_m"] + chunk[-1]["act_pr_m"] + chunk[-1]["act_el_m"]
        + chunk[-1]["act_es_a"] + chunk[-1]["act_pr_a"] + chunk[-1]["act_el_a"],
    }


def main():
    cfg = load_inputs(XLSX)
    rows = simulate(cfg)
    print("VAT comparison — base case simulation (v4 inputs, 10/20/30 visits/yr)")
    print(f"Blended Essential monthly £{cfg['ess_m']:.2f}, provider Essential £{cfg['pay_ess']:.2f}/mo")
    print()
    for y in (1, 2, 5):
        t = year_totals(rows, y)
        d_gp = t["gp_a"] - t["gp_p"]
        d_pr = t["profit_a"] - t["profit_p"]
        pct = 100 * d_gp / t["gp_p"] if t["gp_p"] else 0
        print(f"=== Year {y} ===")
        print(f"  End customers (approx): {t['cust']:.0f}")
        print(f"  Gross customer revenue: £{t['gross']:,.0f}")
        print(f"  Provider pay:           £{t['prov']:,.0f}")
        print(f"  VAT registered — principal: {t['vat_reg_p']} | agent threshold: {t['vat_reg_a']}")
        print(f"  VAT accrued — principal: £{t['vat_p']:,.0f} | agent: £{t['vat_a']:,.0f}")
        print(f"  Gross profit (NetRev − Prov):")
        print(f"    Principal: £{t['gp_p']:,.0f}")
        print(f"    Agent:     £{t['gp_a']:,.0f}")
        print(f"    Uplift:    £{d_gp:,.0f} ({pct:.1f}%)")
        print(f"  Net profit (after all opex):")
        print(f"    Principal: £{t['profit_p']:,.0f}")
        print(f"    Agent:     £{t['profit_a']:,.0f}")
        print(f"    Uplift:    £{d_pr:,.0f}")
        print()

    reg_p = next(r["m"] for r in rows if r["vat_reg_p"])
    reg_a = next((r["m"] for r in rows if r["vat_reg_a"]), None)
    print(f"First month VAT registered — principal: month {reg_p}")
    print(f"First month VAT registered — agent (margin threshold): month {reg_a or 'never in 60 mo'}")


if __name__ == "__main__":
    main()
